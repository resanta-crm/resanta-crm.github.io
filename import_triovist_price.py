#!/usr/bin/env python3
"""RESANTA CRM v23.6.6 — automatic Vitebsk price import for Triovist.

Source: 1C email attachment "Прайс для СРМ (XLSX).xlsx".
Truth price: report parameter "Вид цены: Отпускная" + column "Мелкий опт 2 с НДС".
Duplicate series rows for one SKU are collapsed only when the price is identical.
"Деление на 0" and empty/non-positive prices are ignored.
"""
from __future__ import annotations
import email, imaplib, io, os, re, socket, time, uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from email.header import decode_header
from email.utils import parsedate_to_datetime
import openpyxl
import requests

VERSION='v23.6.6'
IMAP_HOST='imap.gmail.com'; IMAP_PORT=993
IMAP_USER=os.environ['IMAP_USER'].strip()
IMAP_PASS=''.join(os.environ['IMAP_PASS'].split()).replace('\xa0','')
SUPABASE_URL=os.environ['SUPABASE_URL'].strip().rstrip('/')
SUPABASE_KEY=os.environ['SUPABASE_KEY'].strip()
LOOKBACK=max(2,int(os.getenv('TRIOVIST_PRICE_LOOKBACK_DAYS','14')))
MAIL_TIMEOUT=max(20,int(os.getenv('TRIOVIST_PRICE_IMAP_TIMEOUT_SECONDS','45')))
SKU_RE=re.compile(r'^\d+(?:/\d+){1,6}$')
PERIOD_RE=re.compile(r'(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})')


def log(x): print(x,flush=True)
def norm(v): return re.sub(r'\s+',' ',str(v or '').strip().lower().replace('ё','е'))
def compact(v): return re.sub(r'[^a-zа-я0-9]+','',norm(v))
def decode(v):
    if not v:return ''
    out=[]
    for part,enc in decode_header(v):
        if isinstance(part,bytes):
            try: out.append(part.decode(enc or 'utf-8','replace'))
            except Exception: out.append(part.decode('utf-8','replace'))
        else: out.append(part)
    return ''.join(out)
def msg_date(v):
    try:
        d=parsedate_to_datetime(v); return (d.replace(tzinfo=timezone.utc) if d.tzinfo is None else d).astimezone(timezone.utc)
    except Exception:return datetime.min.replace(tzinfo=timezone.utc)
def dec(v):
    if v in (None,''): return None
    s=str(v).replace('\xa0','').replace(' ','').replace(',','.').strip()
    if 'деление' in s.lower(): return None
    try:
        x=Decimal(s)
        return x if x>0 else None
    except (InvalidOperation,ValueError): return None

def connect_mail():
    last=None
    for attempt in range(1,4):
        m=None
        try:
            log(f'{VERSION}: Gmail IMAP {attempt}/3')
            m=imaplib.IMAP4_SSL(IMAP_HOST,IMAP_PORT,timeout=MAIL_TIMEOUT); m.login(IMAP_USER,IMAP_PASS)
            try:m.sock.settimeout(MAIL_TIMEOUT)
            except Exception:pass
            return m
        except Exception as e:
            last=e
            try:
                if m:m.logout()
            except Exception:pass
            if attempt<3:time.sleep(6)
    raise RuntimeError(f'Не удалось подключиться к Gmail: {last}')

def select_box(mail):
    try:
        st,boxes=mail.list()
        if st=='OK':
            for raw in boxes or []:
                if b'\\All' not in raw: continue
                mt=re.search(br'\)\s+"[^"]*"\s+(.+)$',raw)
                if not mt:continue
                token=mt.group(1).strip().strip(b'"').decode('ascii','ignore')
                if token and mail.select(f'"{token}"' if ' ' in token else token)[0]=='OK': return token
    except Exception:pass
    if mail.select('INBOX')[0]!='OK': raise RuntimeError('Не удалось открыть Gmail')
    return 'INBOX'

def excel_attachments(msg):
    out=[]
    for p in msg.walk():
        fn=decode(p.get_filename()).strip()
        if not fn.lower().endswith(('.xlsx','.xlsm')):continue
        body=p.get_payload(decode=True)
        if body:out.append((fn,body))
    return out

def parse_report(blob):
    wb=openpyxl.load_workbook(io.BytesIO(blob),data_only=True,read_only=True)
    ws=wb[wb.sheetnames[0]]
    top=[]
    for row in ws.iter_rows(min_row=1,max_row=min(22,ws.max_row),values_only=True): top.append(list(row))
    joined=' | '.join(str(x or '') for r in top for x in r)
    if 'вид цены: отпускная' not in norm(joined): raise ValueError('Не подтвержден параметр «Вид цены: Отпускная»')
    period=PERIOD_RE.search(joined)
    report_start=report_end=None
    if period:
        report_start=datetime.strptime(period.group(1),'%d.%m.%Y').date().isoformat()
        report_end=datetime.strptime(period.group(2),'%d.%m.%Y').date().isoformat()
    article_col=product_col=price_col=None; header_row=1
    for ri,r in enumerate(top,1):
        for ci,v in enumerate(r,1):
            c=compact(v)
            if c=='артикул':article_col=ci; header_row=max(header_row,ri)
            elif c=='номенклатура':product_col=ci; header_row=max(header_row,ri)
            elif 'мелкийопт2сндс' in c:price_col=ci
    if not article_col or not product_col or not price_col: raise ValueError('Не найдены Артикул / Номенклатура / Мелкий опт 2 с НДС')
    items={}
    for row in ws.iter_rows(min_row=header_row+1,values_only=True):
        sku=str(row[article_col-1] or '').strip().replace("'",'')
        if not SKU_RE.match(sku):continue
        product=str(row[product_col-1] or '').strip()
        price=dec(row[price_col-1])
        if not product or price is None:continue
        old=items.get(sku)
        if old and abs(old['price']-price)>Decimal('0.01'): raise ValueError(f'У артикула {sku} две разные цены: {old["price"]} и {price}')
        items[sku]={'sku':sku,'product':product,'price':price}
    if len(items)<250: raise ValueError(f'Слишком мало корректных SKU: {len(items)}')
    return report_start,report_end,list(items.values())

def headers(): return {'apikey':SUPABASE_KEY,'Authorization':'Bearer '+SUPABASE_KEY,'Content-Type':'application/json','Prefer':'return=minimal'}
def rest(path,method='GET',payload=None,params=None):
    r=requests.request(method,SUPABASE_URL+'/rest/v1/'+path,headers=headers(),json=payload,params=params,timeout=90)
    if r.status_code>=300: raise RuntimeError(f'Supabase {method} {path}: {r.status_code} {r.text[:800]}')
    if not r.text:return None
    try:return r.json()
    except Exception:return None

def already(message_id):
    r=rest('triovist_price_imports',params={'source_message_id':'eq.'+message_id,'select':'id','limit':'1'})
    return bool(r)

def import_report(message_id,subject,filename,report_start,report_end,items):
    iid=str(uuid.uuid4())
    rest('triovist_price_imports','POST',[{'id':iid,'source_message_id':message_id,'source_file':filename,'subject':subject,'report_start':report_start,'report_end':report_end,'row_count':len(items),'status':'success'}])
    try:
        for i in range(0,len(items),300):
            batch=[{'import_id':iid,'sku':x['sku'],'product':x['product'],'price':str(x['price'].quantize(Decimal('0.01')))} for x in items[i:i+300]]
            rest('triovist_price_items','POST',batch)
    except Exception:
        try:rest('triovist_price_imports?id=eq.'+iid,'DELETE')
        except Exception:pass
        raise
    return iid

def main():
    mail=connect_mail(); box=select_box(mail); log('Ищу прайс в '+box)
    try:
        since=(datetime.now()-__import__('datetime').timedelta(days=LOOKBACK)).strftime('%d-%b-%Y')
        st,data=mail.uid('search',None,'SINCE',since)
        if st!='OK': raise RuntimeError('Gmail search failed')
        uids=(data[0] or b'').split()[-120:]
        candidates=[]
        for uid in reversed(uids):
            st,h=mail.uid('fetch',uid,'(BODY.PEEK[HEADER.FIELDS (SUBJECT FROM DATE MESSAGE-ID)])')
            if st!='OK' or not h or not h[0]:continue
            raw=h[0][1] if isinstance(h[0],tuple) else b''; msg=email.message_from_bytes(raw)
            subject=decode(msg.get('Subject'))
            score=(50 if 'прайс для срм' in norm(subject) else 45 if 'прайс для crm' in norm(subject) else 0)
            if not score:continue
            candidates.append((score,msg_date(msg.get('Date')),uid,subject))
        candidates.sort(reverse=True)
        for _,_,uid,subject in candidates[:20]:
            st,full=mail.uid('fetch',uid,'(RFC822)')
            if st!='OK' or not full:continue
            blob=next((x[1] for x in full if isinstance(x,tuple)),None)
            if not blob:continue
            msg=email.message_from_bytes(blob); mid=(msg.get('Message-ID') or uid.decode()).strip('<>')
            if already(mid): log('Уже импортирован: '+subject); return
            for filename,payload in excel_attachments(msg):
                try:
                    rs,re_,items=parse_report(payload)
                    iid=import_report(mid,subject,filename,rs,re_,items)
                    log(f'✅ Прайс импортирован: {len(items)} SKU, {filename}, import={iid}')
                    return
                except Exception as e: log(f'  Не подошёл {filename}: {e}')
        log('Нового валидного прайса не найдено. Данные CRM не менялись.')
    finally:
        try:mail.logout()
        except Exception:pass
if __name__=='__main__':main()
