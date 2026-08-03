#!/usr/bin/env python3
"""Ручная загрузка остатков 21vek и Чехова из Excel в защищённые таблицы Supabase."""
from __future__ import annotations

import hashlib
import os
import re
import sys
import time
import uuid
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import openpyxl
import requests

SUPABASE_URL = os.environ["SUPABASE_URL"].strip().rstrip("/")
SUPABASE_KEY = os.environ["SUPABASE_KEY"].strip()
ARTICLE_RE = re.compile(r"\d+(?:/\d+){1,6}")


def log(message):
    print(message, flush=True)


def dec(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except InvalidOperation as exc:
        raise RuntimeError(f"Некорректное число: {value!r}") from exc


def n3(value):
    return format(dec(value).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP), "f")


def norm(value):
    return re.sub(r"[^a-zа-я0-9]+", "", str(value or "").lower().replace("ё", "е"))


def extract_internal_sku(product):
    matches = ARTICLE_RE.findall(str(product or ""))
    if matches:
        return matches[-1]
    key = norm(product)
    mappings = {
        "асн3000н1с": "63/6/21", "асн3000н1ц": "63/6/21", "asn3000n1c": "63/6/21",
        "clm36li": "70/4/10", "elm1800t": "70/4/5", "dt9208a": "61/10/507",
    }
    for marker, sku in mappings.items():
        if marker in key:
            return sku
    return ""


def parse_partner(path):
    """Разбор остатков 21vek с плавающей строкой шапки.

    Колонки продаж не обязательны: свежий файл остатков может их не содержать.
    В таком случае отправляются пустые метки, а SQL сохраняет предыдущие
    продажи 21vek и обновляет только остатки/заказы.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    raw_rows = ws.iter_rows(values_only=True)
    header_row = None
    header_index = -1
    buffered = []
    for idx, values in enumerate(raw_rows):
        vals = list(values)
        normalized = [norm(v) for v in vals]
        has_article = any(v == "артикул" or v.startswith("артикул") for v in normalized)
        has_product = any(v == "номенклатура" or v.startswith("номенклатура") for v in normalized)
        has_total = any(v == "всего" or v.startswith("всего") for v in normalized)
        if has_article and has_product and has_total:
            header_row = [str(v or "").strip() for v in vals]
            header_index = idx
            break
        if idx >= 29:
            break
    if header_row is None:
        raise RuntimeError("В остатках 21vek не найдена строка шапки с Артикулом, Номенклатурой и Всего")

    headers = header_row
    index = {norm(name): i for i, name in enumerate(headers)}
    def col_any(*labels):
        for label in labels:
            key = norm(label)
            for h, c in index.items():
                if h == key or h.startswith(key):
                    return c
        raise RuntimeError(f"В остатках 21vek нет колонки: {' / '.join(labels)}")

    c_partner = col_any("Артикул", "Артикул 21vek", "Код товара")
    c_kind = col_any("Вид", "Категория")
    c_brand = col_any("Производитель", "Бренд")
    c_product = col_any("Номенклатура", "Наименование")
    c_total = col_any("Всего", "Остаток всего")
    c_free = col_any("Свободно", "Доступно")
    c_transit = col_any("Свободно в пути и резерве", "В пути и резерве", "Резерв")
    c_orders = col_any("Заказы", "В заказах")

    sales_cols = [i for i, name in enumerate(headers) if norm(name).startswith("продажиза")]
    sales_cols = sales_cols[-3:] if len(sales_cols) >= 3 else []
    if not sales_cols:
        log("21vek: в свежем файле нет 3 колонок продаж — обновляются остатки, предыдущие продажи будут сохранены")

    rows, unmatched = [], []
    for idx, values in enumerate(ws.iter_rows(values_only=True)):
        if idx <= header_index:
            continue
        values = list(values)
        if len(values) <= max(c_product, c_partner, c_total):
            continue
        product = str(values[c_product] or "").strip()
        if not product:
            continue
        sku = extract_internal_sku(product)
        if not sku:
            unmatched.append(product)
            continue
        partner_sku = str(values[c_partner] or "").strip()
        kind = str(values[c_kind] or "").strip()
        excluded = "+" in product or "комплект" in norm(kind + " " + product) or "невыводить" in norm(product)
        row_key = hashlib.md5(f"{partner_sku}|{sku}|{product}".encode("utf-8")).hexdigest()
        sales_values = [n3(values[c]) for c in sales_cols] if sales_cols else ["0.000", "0.000", "0.000"]
        sales_labels = [headers[c].replace("Продажи за", "").strip() for c in sales_cols] if sales_cols else [None, None, None]
        rows.append({
            "row_key": row_key, "partner_sku": partner_sku, "sku": sku,
            "kind": kind, "brand": str(values[c_brand] or "").strip(), "product": product,
            "sales_m1": sales_values[0], "sales_m2": sales_values[1], "sales_m3": sales_values[2],
            "sales_m1_label": sales_labels[0], "sales_m2_label": sales_labels[1], "sales_m3_label": sales_labels[2],
            "qty_total": n3(values[c_total]), "qty_free": n3(values[c_free]),
            "qty_transit_reserve": n3(values[c_transit]), "qty_orders": n3(values[c_orders]),
            "excluded": excluded,
            "match_note": "Исключено из рекомендации: комплект" if excluded else ("Продажи сохранены из предыдущей загрузки" if not sales_cols else None),
        })
    if unmatched:
        raise RuntimeError("Не сопоставлены артикулы 21vek: " + "; ".join(unmatched[:10]))
    if len(rows) < 100:
        raise RuntimeError(f"Слишком мало строк 21vek: {len(rows)}")
    return rows

def parse_chekhov(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    all_rows = ws.iter_rows(values_only=True)
    headers = None
    data_rows = []
    for idx, values in enumerate(all_rows, start=1):
        vals = list(values)
        if headers is None:
            normalized = [norm(v) for v in vals]
            if "номенклатура" in normalized and "артикул" in normalized:
                headers = [str(v or "").strip() for v in vals]
            continue
        data_rows.append(vals)
    if headers is None:
        raise RuntimeError("В остатках Чехова не найдена шапка")
    def find_col(label):
        key = norm(label)
        for i, value in enumerate(headers):
            if norm(value) == key or norm(value).startswith(key):
                return i
        raise RuntimeError(f"В остатках Чехова нет колонки «{label}»")
    c_product, c_sku, c_box = find_col("Номенклатура"), find_col("Артикул"), find_col("Количество в коробке")
    c_stock, c_near, c_total = find_col("Остаток"), find_col("Остаток рядом"), find_col("Сумма")
    rows, seen = [], set()
    for values in data_rows:
        sku = str(values[c_sku] or "").strip().lstrip("'")
        if not re.fullmatch(r"\d+(?:/\d+){1,6}", sku):
            continue
        if sku in seen:
            raise RuntimeError(f"Дубль артикула Чехова: {sku}")
        seen.add(sku)
        product = str(values[c_product] or "").strip()
        rows.append({
            "sku": sku, "product": product, "box_qty": n3(values[c_box]),
            "qty_stock": n3(values[c_stock]), "qty_nearby": n3(values[c_near]),
            "qty_total": n3(values[c_total]),
        })
    if len(rows) < 100:
        raise RuntimeError(f"Слишком мало строк Чехова: {len(rows)}")
    return rows


def rpc(name, payload, timeout=90, attempts=3):
    url = f"{SUPABASE_URL}/rest/v1/rpc/{name}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=timeout)
            if response.status_code in (200, 201, 204):
                return response.json() if response.text.strip() else {"ok": True}
            last_error = RuntimeError(f"RPC {name}: {response.status_code} {response.text}")
        except requests.RequestException as exc:
            last_error = exc
        if attempt < attempts:
            wait = attempt * 3
            log(f"Повтор RPC {name} через {wait} сек. Попытка {attempt + 1}/{attempts}")
            time.sleep(wait)
    raise RuntimeError(f"Supabase RPC {name} не выполнен: {last_error}")


def upload_staged(source, rows, filename, batch_size=300):
    import_id = uuid.uuid4().hex
    rpc("triovist_stock_stage_begin", {
        "p_source": source,
        "p_import_id": import_id,
    })
    try:
        total = len(rows)
        for offset in range(0, total, batch_size):
            batch = rows[offset:offset + batch_size]
            result = rpc("triovist_stock_stage_batch", {
                "p_source": source,
                "p_import_id": import_id,
                "p_rows": batch,
            })
            loaded = min(offset + len(batch), total)
            log(f"{source}: загружено {loaded}/{total} строк; staging={result.get('staged_rows')}")
        return rpc("triovist_stock_stage_commit", {
            "p_source": source,
            "p_import_id": import_id,
            "p_expected_rows": total,
            "p_source_file": filename,
            "p_snapshot_date": date.today().isoformat(),
        }, timeout=180)
    except Exception:
        try:
            rpc("triovist_stock_stage_abort", {
                "p_source": source,
                "p_import_id": import_id,
            }, timeout=30, attempts=1)
        except Exception as abort_error:
            log(f"Не удалось очистить staging {source}: {abort_error}")
        raise


def main():
    partner = Path(sys.argv[1] if len(sys.argv) > 1 else "data/stock_21vek.xlsx")
    chekhov = Path(sys.argv[2] if len(sys.argv) > 2 else "data/stock_chekhov.xlsx")
    if not partner.exists() or not chekhov.exists():
        raise RuntimeError(f"Не найдены файлы: {partner}, {chekhov}")
    partner_rows = parse_partner(partner)
    chekhov_rows = parse_chekhov(chekhov)
    log(f"21vek: {len(partner_rows)} строк, исключено комплектов: {sum(1 for r in partner_rows if r['excluded'])}")
    log(f"Чехов: {len(chekhov_rows)} уникальных артикулов")
    log(f"21vek RPC: {upload_staged('partner', partner_rows, partner.name)}")
    log(f"Чехов RPC: {upload_staged('chekhov', chekhov_rows, chekhov.name)}")
    log("✅ Остатки 21vek и Чехова загружены")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log(f"ОШИБКА: {exc}")
        sys.exit(1)
