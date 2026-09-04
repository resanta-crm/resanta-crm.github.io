#!/usr/bin/env python3
"""
Импорт ежемесячного отчёта продаж из почты в Supabase (purchase_history).
Версия контроля разбора: v22.7.19 — пакетное чтение IMAP-заголовков + сверка каждого контрагента с итогом 1С.

Как работает:
  1. Заходит на почту по IMAP (Gmail).
  2. Ищет письма за последние LOOKBACK_DAYS дней с темой "Продажи для CRM"
     и берёт САМОЕ СВЕЖЕЕ.
  3. Достаёт вложение .xlsx и разбирает его.
  4. Заменяет в purchase_history данные ЗА МЕСЯЦ ИЗ ОТЧЁТА на новые
     (старые месяцы не трогает — история накапливается, на ней стоит маст-лист).

ДВЕ ИСПРАВЛЕННЫЕ ОШИБКИ (из-за них продажи молча не обновлялись):

  1. Поиск писем шёл по флагу UNSEEN — только непрочитанные. Стоило кому-то
     открыть письмо в Gmail, и скрипт переставал его видеть: писал "новых писем
     нет" и завершался УСПЕШНО. Данные в CRM застывали, Action был зелёный.
     Теперь флаг прочтения не участвует: ищем по теме и дате.

  2. replace_month() СНАЧАЛА удалял месяц и только потом смотрел, есть ли что
     писать. Иерархия в отчёте держится на ОТСТУПАХ ячеек — стоит 1С изменить
     оформление, и парсер вернёт 0 строк. Месяц бы стёрся, а маст-лист у всех
     менеджеров опустел. Теперь пустой разбор — это ошибка, и до удаления
     дело не доходит.

Зачем: на основе purchase_history в CRM строится МАСТ-ЛИСТ клиента —
"какие товары этот клиент покупает". Раньше историю грузили вручную.

Структура отчёта 1С ("Валовая прибыль предприятия"):
  Иерархия задана ОТСТУПАМИ в первой колонке:
    отступ 0 — Клиент          (напр. "Руд Буд ЧТУП")
    отступ 2 — Категория       (напр. "САДОВАЯ ТЕХНИКА")
    отступ 4 — Подгруппа       (напр. "Прочее")
    отступ 6 — Товар           (напр. "Опрыскиватель SP-10AC Huter")
  Менеджеры — в КОЛОНКАХ (у каждого своя пара "Количество"/"Выручка"),
  их имена записаны в строке-шапке. Мы находим эти колонки по именам,
  чтобы не зависеть от их порядка и количества.
  Период (месяц) берём из строки "Параметры: Период: 01.07.2026 - 31.07.2026".

Все секреты — из переменных окружения (в GitHub из Secrets).
"""

import os
import sys
import re
import imaplib
import email
import io
import time
from email.header import decode_header
from email.utils import parsedate_to_datetime
from datetime import datetime, date, timedelta, timezone

import openpyxl
import requests

# ---------- Настройки ----------
IMAP_HOST = os.environ.get("IMAP_HOST", "imap.gmail.com")
IMAP_PORT = int(os.environ.get("IMAP_PORT", "993"))
IMAP_USER = os.environ["IMAP_USER"].strip()
# Пароль приложения Google копируется с неразрывными пробелами — вычищаем их.
IMAP_PASS = "".join(os.environ["IMAP_PASS"].split()).replace("\xa0", "")

SUPABASE_URL = os.environ["SUPABASE_URL"].strip().rstrip("/")
SUPABASE_KEY = os.environ["SUPABASE_KEY"].strip()

SUBJECT_MARKER = "Продажи для CRM"

# За сколько дней искать письма. Отчёт по продажам приходит ежедневно в 10:00.
LOOKBACK_DAYS = int(os.environ.get("SALES_LOOKBACK_DAYS", "14"))
MAIL_TIMEOUT = int(os.environ.get("SALES_IMAP_TIMEOUT_SECONDS", "45"))
MAX_FULL_MESSAGES = int(os.environ.get("SALES_MAX_EMAIL_CANDIDATES", "8"))
HEADER_BATCH_SIZE = max(10, int(os.environ.get("SALES_IMAP_HEADER_BATCH_SIZE", "75")))
HEADER_SCAN_DEADLINE = max(60, int(os.environ.get("SALES_IMAP_HEADER_SCAN_DEADLINE_SECONDS", "180")))

# Менеджеры, чьи продажи грузим. Остальных (Савон, Азаров и др.) игнорируем.
ALLOWED_MANAGERS = ["Руднев", "Ачинович", "Шкуран"]

# Уровни иерархии по отступу в Excel.
LEVEL_CLIENT, LEVEL_CATEGORY, LEVEL_SUBGROUP, LEVEL_PRODUCT = 0, 2, 4, 6

MONTHS_RU = {
    1: "янв", 2: "фев", 3: "мар", 4: "апр", 5: "май", 6: "июн",
    7: "июл", 8: "авг", 9: "сен", 10: "окт", 11: "ноя", 12: "дек",
}


def log(msg):
    print(msg, flush=True)



IMPORT_CONTEXT = {"report_period": None, "report_date": None, "source_message_at": None, "row_count": None}


def _api_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }


def set_import_status(source, status, *, report_period=None, report_date=None,
                      source_message_at=None, row_count=None, details=None,
                      error_text=None):
    """Записывает фактический статус импорта в Supabase.

    Ошибка этой служебной записи не должна уничтожать основной импорт, поэтому
    вызывающая сторона использует safe_set_import_status().
    """
    payload = {
        "p_source": source,
        "p_status": status,
        "p_report_period": report_period,
        "p_report_date": report_date,
        "p_source_message_at": source_message_at,
        "p_row_count": row_count,
        "p_details": details,
        "p_error_text": error_text,
    }
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/rpc/crm_set_import_status",
        headers=_api_headers(), json=payload, timeout=30,
    )
    if resp.status_code not in (200, 201, 204):
        raise RuntimeError(f"Не удалось записать статус импорта: {resp.status_code} {resp.text}")


def safe_set_import_status(source, status, **kwargs):
    try:
        set_import_status(source, status, **kwargs)
    except Exception as exc:
        log(f"  ⚠️ Статус импорта не записан: {exc}")



def sales_message_already_loaded(report_period, sent):
    """Hourly guard: the same 1C email must not rewrite purchase_history again."""
    try:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/crm_import_status",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
            params={
                "source": "eq.sales",
                "select": "status,report_period,source_message_at",
                "limit": "1",
            },
            timeout=20,
        )
        if resp.status_code != 200:
            return False
        rows = resp.json() or []
        if not rows:
            return False
        row = rows[0]
        if str(row.get("status") or "") != "ok" or str(row.get("report_period") or "") != str(report_period or ""):
            return False
        raw = str(row.get("source_message_at") or "").strip()
        if not raw:
            return False
        prev = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        if prev.tzinfo is None:
            prev = prev.replace(tzinfo=timezone.utc)
        cur = sent if sent.tzinfo else sent.replace(tzinfo=timezone.utc)
        return abs((prev.astimezone(timezone.utc) - cur.astimezone(timezone.utc)).total_seconds()) < 1
    except Exception as exc:
        log(f"  ⚠️ Не удалось проверить дубль письма: {exc}")
        return False


def capture_promotion_sales_snapshots(report_month, source_message_at):
    """Capture action sales only when a genuinely new sales email was imported."""
    try:
        resp = requests.post(
            f"{SUPABASE_URL}/rest/v1/rpc/promotion_capture_sales_snapshots_v23654",
            headers=_api_headers(),
            json={
                "p_report_month": report_month,
                "p_source_message_at": source_message_at,
            },
            timeout=60,
        )
        if resp.status_code not in (200, 201, 204):
            log(f"  ⚠️ Срез продаж по акциям не снят: {resp.status_code} {resp.text[:300]}")
            return False
        log("  ✅ Срез продаж по действующим акциям обновлён.")
        return True
    except Exception as exc:
        log(f"  ⚠️ Срез продаж по акциям не снят: {exc}")
        return False


def decode_mime_header(value):
    if not value:
        return ""
    parts = []
    for part, encoding in decode_header(value):
        if not isinstance(part, bytes):
            parts.append(part)
            continue
        enc = encoding or "utf-8"
        if str(enc).lower() == "unknown-8bit":
            enc = "utf-8"
        try:
            parts.append(part.decode(enc, errors="replace"))
        except (LookupError, UnicodeDecodeError):
            parts.append(part.decode("utf-8", errors="replace"))
    return "".join(parts)


def normalize_message_date(value):
    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return datetime.min.replace(tzinfo=timezone.utc)


def select_mailbox(mail):
    """Ищет папку Gmail с флагом \\All, чтобы импорт видел даже архивированные
    фильтром письма. Если специальная папка недоступна, использует INBOX.
    """
    try:
        status, boxes = mail.list()
        if status == "OK":
            for raw in boxes or []:
                if b"\\All" not in raw:
                    continue
                match = re.search(br'\)\s+"[^"]*"\s+(.+)$', raw)
                if not match:
                    continue
                token = match.group(1).strip()
                if token.startswith(b'"') and token.endswith(b'"'):
                    token = token[1:-1]
                mailbox = token.decode("ascii", errors="ignore")
                if mailbox:
                    select_arg = f'"{mailbox}"' if " " in mailbox else mailbox
                    st, _ = mail.select(select_arg)
                    if st == "OK":
                        log(f"Ищу письма во всей почте: {mailbox}")
                        return mailbox
    except Exception as exc:
        log(f"  ⚠️ Не удалось открыть папку всей почты: {exc}")
    st, _ = mail.select("INBOX")
    if st != "OK":
        raise RuntimeError("Не удалось открыть INBOX")
    log("Ищу письма в INBOX")
    return "INBOX"


def fetch_message_headers_batch(mail, msg_ids):
    """Одним IMAP FETCH получает Subject/Date сразу для пачки писем.

    В v22.7.18 каждый из ~200 message-id запрашивался отдельно и Gmail мог
    тратить 10–20 минут только на заголовки. Здесь на 213 писем будет 3 FETCH
    при размере пачки 75. Полные письма и XLSX на этом этапе не скачиваются.
    """
    if not msg_ids:
        return []
    message_set = b",".join(msg_ids)
    status, data = mail.fetch(message_set, "(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE)])")
    if status != "OK":
        raise RuntimeError("Gmail не вернул пакет заголовков")

    result = []
    requested = {bytes(x) for x in msg_ids}
    for item in data or []:
        if not (isinstance(item, tuple) and len(item) > 1 and isinstance(item[1], (bytes, bytearray))):
            continue
        meta_raw = item[0] if isinstance(item[0], (bytes, bytearray)) else b""
        m = re.match(rb"\s*(\d+)", bytes(meta_raw))
        if not m:
            continue
        msg_id = m.group(1)
        # Gmail обычно возвращает sequence number из нашего message-set.
        # Защита нужна, чтобы случайная служебная строка ответа не стала письмом.
        if msg_id not in requested:
            continue
        raw = bytes(item[1])
        if not raw:
            continue
        msg = email.message_from_bytes(raw)
        result.append({
            "id": msg_id,
            "subject": decode_mime_header(msg.get("Subject", "")),
            "sent": normalize_message_date(msg.get("Date")),
        })
    return result


def sales_mail_candidates(mail, ids):
    """Пакетно читает только Subject/Date и возвращает свежие письма 1С.

    Никакого цикла из сотен отдельных IMAP FETCH. Есть общий дедлайн фазы,
    прогресс по пачкам и ограничение числа полных писем, которые будут скачаны.
    """
    out = []
    started = time.monotonic()
    total_batches = max(1, (len(ids) + HEADER_BATCH_SIZE - 1) // HEADER_BATCH_SIZE)
    for batch_no, start in enumerate(range(0, len(ids), HEADER_BATCH_SIZE), 1):
        if time.monotonic() - started > HEADER_SCAN_DEADLINE:
            raise RuntimeError(
                f"Gmail слишком долго отдаёт заголовки: превышен лимит {HEADER_SCAN_DEADLINE} сек. "
                "Импорт остановлен без изменения данных CRM."
            )
        batch = ids[start:start + HEADER_BATCH_SIZE]
        log(f"  Заголовки {batch_no}/{total_batches}: {len(batch)} писем...")
        try:
            metas = fetch_message_headers_batch(mail, batch)
        except Exception as exc:
            raise RuntimeError(f"Не удалось получить пакет заголовков {batch_no}/{total_batches}: {exc}") from exc
        for meta in metas:
            if SUBJECT_MARKER.lower() in meta["subject"].lower():
                out.append(meta)

    out.sort(key=lambda x: x["sent"], reverse=True)
    log(f"Найдено писем «{SUBJECT_MARKER}»: {len(out)}.")
    if len(out) > MAX_FULL_MESSAGES:
        log(f"Полностью загружу максимум {MAX_FULL_MESSAGES} самых свежих; как только найден корректный отчёт текущего месяца — остановлюсь.")
    elif out:
        log("Проверяю вложение самого свежего письма; остальные — только запасные.")
    return out[:MAX_FULL_MESSAGES]


def short_manager(full_name):
    """'Руднев Александр Александрович' -> 'Руднев' (как в CRM)."""
    for surname in ALLOWED_MANAGERS:
        if surname.lower() in (full_name or "").lower():
            return surname
    return None


def parse_period(ws):
    """Достаёт месяц отчёта из строки 'Период: 01.07.2026 - 31.07.2026'.
    Возвращает '2026-07-01' — как хранится month в purchase_history."""
    for row in ws.iter_rows(min_row=1, max_row=12, max_col=6):
        for cell in row:
            if not cell.value:
                continue
            m = re.search(r"Период:\s*(\d{2})\.(\d{2})\.(\d{4})", str(cell.value))
            if m:
                day, month, year = m.groups()
                return f"{year}-{month}-01"
    raise RuntimeError("Не нашёл период в отчёте (строка 'Период: ...')")



def parse_report_bounds(xlsx_bytes):
    """Возвращает (start_date, end_date) периода отчёта 1С.

    Для оперативного YoY важно знать не только месяц, но и точную конечную
    дату отчёта: 01–07.08.2026 надо сравнивать с 01–07.08.2025, а не с полным
    августом прошлого года.
    """
    if xlsx_bytes[:2] != b"PK":
        raise RuntimeError("Исторический снимок продаж должен быть в формате .xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.active
    rx = re.compile(
        r"Период:\s*(\d{2})\.(\d{2})\.(\d{4})\s*[-–—]\s*(\d{2})\.(\d{2})\.(\d{4})",
        re.I,
    )
    for row in ws.iter_rows(min_row=1, max_row=12, max_col=8, values_only=True):
        for value in row:
            if not value:
                continue
            m = rx.search(str(value))
            if m:
                d1, m1, y1, d2, m2, y2 = map(int, m.groups())
                return date(y1, m1, d1), date(y2, m2, d2)
    raise RuntimeError("Не нашёл полный период отчёта 1С (дата начала и дата окончания)")


def _previous_year_same_date(value):
    try:
        return value.replace(year=value.year - 1)
    except ValueError:  # 29 февраля -> 28 февраля
        return value.replace(year=value.year - 1, day=28)

def find_manager_columns(ws):
    """Находит, в каких колонках лежат данные каждого менеджера.

    В шапке отчёта имена менеджеров стоят над их парой колонок
    'Количество' / 'Выручка'. Ищем имена, затем под ними — заголовки колонок.
    Возвращает {'Руднев': {'qty': 13, 'revenue': 14}, ...} (1-based).
    """
    managers = {}
    header_row = None

    for row_idx in range(1, 15):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if not val:
                continue
            short = short_manager(str(val))
            if short:
                managers[short] = {"name_col": col_idx}
                header_row = row_idx

    if not managers:
        raise RuntimeError("Не нашёл в отчёте ни одного из менеджеров: " + ", ".join(ALLOWED_MANAGERS))

    # Под строкой с именами идёт строка с 'Количество' / 'Выручка'.
    sub_row = header_row + 1
    for short, info in managers.items():
        c = info["name_col"]
        qty_col = rev_col = None
        # Имя менеджера стоит над первой из его двух колонок.
        for offset in (0, 1):
            label = ws.cell(row=sub_row, column=c + offset).value
            if not label:
                continue
            label = str(label).strip().lower()
            if label.startswith("кол"):
                qty_col = c + offset
            elif label.startswith("выруч"):
                rev_col = c + offset
        if qty_col is None or rev_col is None:
            raise RuntimeError(f"Не нашёл колонки Количество/Выручка для менеджера {short}")
        info["qty"] = qty_col
        info["revenue"] = rev_col

    return managers


def num(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


import re as _re

# Артикул в отчётах Ресанты — числа через слэш: "70/1/67", "73/7/2/32", "64/40".
# Число частей разное (2-4), поэтому шаблон — просто цифры и слэши.
_ARTICLE_RE = _re.compile(r"^\d+(?:/\d+){1,4}$")


def extract_article(value):
    """Возвращает чистый артикул из строки 1С.

    Поддерживает обычный код, апостроф Excel и подпись «Артикул:». Модель
    товара (например ДП-190/1800) артикулом не считается: допустимы только
    цифры, разделённые слэшами.
    """
    text = str(value or "").strip().lstrip("'’` ")
    text = _re.sub(r"^артикул\s*:?\s*", "", text, flags=_re.I).strip()
    return text if _ARTICLE_RE.fullmatch(text) else None


def validate_sku_rows(rows):
    """Не позволяет заменить месяц строками без корректного артикула."""
    bad = [r for r in rows if not extract_article(r.get("sku"))]
    if bad:
        sample = "; ".join(str(r.get("product") or "без названия") for r in bad[:5])
        raise RuntimeError(
            f"В разобранных продажах {len(bad)} строк без корректного артикула. "
            f"Месяц в CRM НЕ заменён. Примеры: {sample}"
        )


def _row_has_manager_turnover(ws, row_idx, managers):
    """Есть ли в строке итог по хотя бы одному полевому менеджеру."""
    for cols in managers.values():
        rev = num(ws.cell(row=row_idx, column=cols["revenue"]).value)
        qty = num(ws.cell(row=row_idx, column=cols["qty"]).value)
        if (rev is not None and abs(rev) > 1e-9) or (qty is not None and abs(qty) > 1e-9):
            return True
    return False


def validate_client_totals(ws, managers, rows, client_header_rows):
    """Контроль суммы КАЖДОГО контрагента, а не только менеджера целиком.

    Старый контроль мог пропустить ситуацию, когда часть SKU одного клиента
    из-за уровня отступа была приписана соседнему клиенту: общий итог менеджера
    при этом всё равно сходился. Теперь месяц не заменяется, пока сумма по
    каждому контрагенту не совпадёт с его строкой-итогом в 1С.
    """
    if not client_header_rows:
        raise RuntimeError("Не удалось определить строки итогов контрагентов в отчёте 1С")

    expected = {}
    for client, row_idx in client_header_rows:
        for mgr, cols in managers.items():
            value = num(ws.cell(row=row_idx, column=cols["revenue"]).value)
            if value is None:
                continue
            expected[(client, mgr)] = expected.get((client, mgr), 0.0) + float(value)

    actual = {}
    for row in rows:
        key = (row["client_name"], row["manager_name"])
        actual[key] = actual.get(key, 0.0) + float(row.get("revenue") or 0.0)

    errors = []
    for (client, mgr), exp in expected.items():
        # Нулевые строки менеджера не требуют дочерней детализации.
        if abs(exp) <= 1e-9:
            continue
        got = actual.get((client, mgr), 0.0)
        tolerance = max(0.05, abs(exp) * 0.0005)
        if abs(got - exp) > tolerance:
            errors.append(f"{client} / {mgr}: в строке 1С {exp:,.2f}, по SKU разобрано {got:,.2f}")

    if errors:
        sample = "; ".join(errors[:20])
        raise RuntimeError(
            "Контроль по контрагентам не пройден. Месяц в CRM НЕ заменён. "
            "Это защищает от ложных падающих клиентов. " + sample
        )
    log(f"  Контроль по контрагентам: проверено {len(client_header_rows)} строк — суммы совпали с SKU.")


def validate_report_totals(ws, managers, rows):
    """Сверяет сумму распознанных SKU с итогом отчёта 1С по менеджерам.

    Если структура отчёта изменилась и часть товаров пропала из разбора,
    импорт останавливается до удаления старого месяца.
    """
    total_row = None
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if str(value or "").strip().lower() == "итого":
            total_row = row_idx
    if total_row is None:
        log("  ⚠️ В отчёте не найдено поле «Итого» — контроль полноты по сумме пропущен.")
        return

    parsed = {}
    for row in rows:
        mgr = row["manager_name"]
        parsed[mgr] = parsed.get(mgr, 0.0) + float(row.get("revenue") or 0)

    errors = []
    for mgr, cols in managers.items():
        expected = num(ws.cell(row=total_row, column=cols["revenue"]).value)
        if expected is None:
            continue
        actual = parsed.get(mgr, 0.0)
        tolerance = max(1.0, abs(expected) * 0.001)
        if abs(actual - expected) > tolerance:
            errors.append(f"{mgr}: в 1С {expected:,.2f}, по SKU разобрано {actual:,.2f}")

    if errors:
        raise RuntimeError(
            "Контроль полноты продаж не пройден. Месяц в CRM НЕ заменён. "
            + "; ".join(errors)
        )
    log("  Контроль полноты: сумма SKU совпадает с итогом отчёта 1С.")


def parse_sales(xlsx_bytes):
    """Разбирает отчёт продаж. Возвращает (month, rows).

    УСТОЙЧИВОСТЬ К ОТСТУПАМ: раньше уровни (клиент/категория/подгруппа/товар)
    определялись по жёстким числам indent (0/2/4/6). Стоило в 1С изменить
    отчёт (добавить артикул) — отступы стали 0/3/6/9/12/15, и парсер перестал
    находить товары. Теперь роль строки определяем ПО СОДЕРЖИМОМУ и по ПОРЯДКУ
    отступов, а не по конкретным числам:
      - самый маленький indent = клиент;
      - строки КАПСОМ (или следующий уровень) = категория;
      - следующий = подгруппа;
      - строка с выручкой и обычным названием = товар;
      - строка вида "70/1/67" = артикул предыдущего товара.
    """
    # Парсер работает с .xlsx (openpyxl). Если рассылка 1С внезапно шлёт старый
    # .xls (начинается с байтов D0CF, а не PK) — не пытаемся гадать, а говорим
    # прямо: переключите формат вложения на .xlsx.
    if xlsx_bytes[:2] != b"PK":
        raise RuntimeError(
            "Отчёт продаж пришёл в старом формате .xls. Скрипт читает .xlsx. "
            "В рассылке 1С выберите формат вложения «Лист Excel 2007-...» (.xlsx)."
        )
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    month = parse_period(ws)
    managers = find_manager_columns(ws)
    log(f"  Месяц отчёта: {month}")
    log(f"  Менеджеры в отчёте: {', '.join(managers.keys())}")

    # Собираем значимые строки с отступами.
    body = []
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        name = cell.value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        if name in ("Контрагент", "Номенклатура", "Параметры:", "Отбор:", "Итого", "Артикул"):
            continue
        body.append((row_idx, name, cell.alignment.indent or 0))

    # НАДЁЖНЫЙ ПРИЗНАК ТОВАРА: строка, СРАЗУ ПОД которой идёт строка-артикул.
    # Раздел (категория/подгруппа/под-подгруппа) артикула под собой не имеет.
    # Это не зависит ни от отступов, ни от глубины иерархии, ни от КАПСа —
    # поэтому переживёт любые изменения структуры отчёта в 1С.
    is_product = [False] * len(body)
    for i in range(len(body) - 1):
        nxt_name = body[i + 1][1]
        if not extract_article(body[i][1]) and extract_article(nxt_name):
            is_product[i] = True

    # Уровень клиента определяем НЕ по самой левой надписи листа (там может быть
    # заголовок отчёта), а по самой левой НЕ-товарной строке, в которой реально
    # есть итог Количество/Выручка менеджера. Это защищает границы контрагентов.
    client_indents = []
    for i, (row_idx, name, indent) in enumerate(body):
        if is_product[i] or extract_article(name):
            continue
        if _row_has_manager_turnover(ws, row_idx, managers):
            client_indents.append(indent)
    lvl_client = min(client_indents) if client_indents else min((ind for _, _, ind in body), default=0)

    rows = []
    client = None
    client_header_rows = []
    # Категорию/подгруппу отслеживаем как "последний не-товар над товаром".
    # Для задач важнее всего категория верхнего уровня — берём самый левый
    # не-товарный заголовок из текущей цепочки.
    heading_stack = {}   # indent -> name

    for i, (row_idx, name, indent) in enumerate(body):
        if extract_article(name):
            continue  # артикулы обрабатываем при товаре

        if is_product[i]:
            if not client:
                continue
            # Категория = самый левый заголовок выше клиента; подгруппа = самый
            # глубокий. Собираем из стека заголовков, что накопился.
            headings = [heading_stack[k] for k in sorted(heading_stack) if k > lvl_client]
            category = headings[0] if headings else "Без категории"
            subgroup = headings[-1] if len(headings) > 1 else (headings[0] if headings else "Прочее")

            # SKU — из следующей строки.
            sku = extract_article(body[i + 1][1]) if i + 1 < len(body) else None

            for mgr, cols in managers.items():
                qty = num(ws.cell(row=row_idx, column=cols["qty"]).value)
                revenue = num(ws.cell(row=row_idx, column=cols["revenue"]).value)
                if not revenue:
                    continue
                rows.append({
                    "client_name": client,
                    "category": category,
                    "subgroup": subgroup,
                    "product": name,
                    "sku": sku,
                    "month": month,
                    "qty": int(qty) if qty is not None else 0,
                    "revenue": revenue,
                    "manager_name": mgr,
                })
            continue

        # Это ЗАГОЛОВОК (не товар, не артикул). Строки левее фактического
        # уровня контрагента — заголовки отчёта, а не клиенты.
        if indent < lvl_client:
            continue
        if indent == lvl_client:
            client = name
            heading_stack = {}   # новый клиент — сбрасываем цепочку разделов
            if _row_has_manager_turnover(ws, row_idx, managers):
                client_header_rows.append((client, row_idx))
        else:
            # Убираем из стека всё, что глубже или равно текущему уровню, и
            # кладём себя — так стек всегда отражает актуальную цепочку.
            for k in [k for k in heading_stack if k >= indent]:
                del heading_stack[k]
            heading_stack[indent] = name

    validate_sku_rows(rows)
    validate_client_totals(ws, managers, rows, client_header_rows)
    validate_report_totals(ws, managers, rows)
    return month, rows


# ===== АВТОЗАВЕДЕНИЕ КЛИЕНТОВ ИЗ 1С =====
# Проблема, которую это решает: клиент заведён в 1С, покупает, его закупки
# лежат в purchase_history — а карточки в CRM нет. Значит он не виден в списке
# клиентов, не попадает в маршруты, в задачи и в ИИ-анализ. Самые свежие
# клиенты были невидимы для менеджеров.
#
# ВАЖНО: заводим ТОЛЬКО из отчёта продаж. В отчёте ПДЗ названия идут как
# "ТТ, ИТ технологии ЧП Могилёвская обл, ..." — это торговые точки, а не
# юрлица; из них мы бы наплодили мусор.

# Юр. формы, которые надо отбросить при сравнении имён. Список тот же, что в
# CRM (normalizeClientName) — иначе "Аникогрупп ЧТПУП" из 1С не совпадёт с
# "Аникогрупп ЧТПУП" из базы и заведётся дубль.
NAME_STOP = {"ооо", "одо", "уп", "чуп", "чтуп", "чпуп", "чтпуп", "ип",
             "оао", "зао", "учп", "чп", "тт", "головной"}


def normalize_name(name):
    """Приводит название к виду, по которому можно сравнивать.

    Кириллица + \b в регэкспах не дружат, поэтому режем по токенам, как в CRM.
    """
    s = str(name or "").strip().lower()
    s = re.sub(r"[«»\"'.,()\-–—]", " ", s)
    tokens = [t for t in s.split() if t and t not in NAME_STOP]
    return " ".join(tokens)


def names_match(a, b):
    """Совпадение имён: точное или по вхождению.

    Вхождение проверяем только для достаточно длинных строк — иначе короткое
    "рудбуд" склеится с чем попало.
    """
    if not a or not b:
        return False
    if a == b:
        return True
    if len(a) >= 6 and len(b) >= 6:
        return a in b or b in a
    return False


ALIAS_RE = re.compile(r"\[\[CRM_ALIASES:([^\]]*)\]\]\s*$")

def client_name_variants(row):
    """Все имена объединённых ТТ, по которым может прийти продажа из 1С."""
    names = [row.get("name") or ""]
    assortment = str(row.get("assortment") or "")
    m = ALIAS_RE.search(assortment)
    if m:
        try:
            import urllib.parse, json
            names.extend(json.loads(urllib.parse.unquote(m.group(1))) or [])
        except Exception:
            pass
    # Поддержка старых объединений, где все названия оставили в одной строке.
    legal = re.compile(r"(?:^|[\s,;/]+)(ООО|ОДО|УП|ЧУП|ЧТУП|ЧПУП|ЧТПУП|ИП|ОАО|ЗАО|УЧП|ЧП)\s+", re.I)
    s = str(row.get("name") or "").strip()
    starts = [m.start(1) for m in legal.finditer(s)]
    if len(starts) > 1:
        for i, start in enumerate(starts):
            part = s[start:(starts[i+1] if i+1 < len(starts) else len(s))].strip(" ,;/")
            if part:
                names.append(part)
    return list(dict.fromkeys(n for n in names if n))

def canonical_client_key(name):
    """Ключ клиента для точной привязки истории продаж к карточке CRM."""
    return re.sub(r"[^a-zа-я0-9]+", "", normalize_name(name).replace("ё", "е"))


def attach_client_ids(rows):
    """Добавляет purchase_history.client_id по текущей карточке и её алиасам.

    Автоматически используем только ОДНОЗНАЧНЫЙ точный ключ. Если одинаковый
    алиас относится к двум карточкам, строка остаётся без client_id и попадает в
    контроль качества данных — так безопаснее, чем приписать продажи не тому.
    """
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }

    clients = _get_all_rest_rows("clients", "id,name,assortment")
    aliases = []
    try:
        aliases = _get_all_rest_rows("client_aliases", "client_id,alias_name")
    except Exception as exc:
        log(f"  ⚠️ Таблица client_aliases недоступна, использую имена из карточек: {exc}")

    keys = {}
    for c in clients:
        cid = str(c.get("id") or "")
        names = client_name_variants(c)
        names.extend(a.get("alias_name") for a in aliases if str(a.get("client_id") or "") == cid)
        for name in names:
            key = canonical_client_key(name)
            if not key:
                continue
            keys.setdefault(key, set()).add(cid)

    unique = {k: next(iter(ids)) for k, ids in keys.items() if len(ids) == 1}
    matched = 0
    unresolved = set()
    ambiguous = set()

    # PostgREST требует одинаковый набор ключей во всех объектах массовой
    # вставки. Поэтому client_id присутствует у КАЖДОЙ строки: UUID либо null.
    # Раньше ключ добавлялся только найденным клиентам, из-за чего импорт падал
    # с PGRST102 «All object keys must match».
    for row in rows:
        row["client_id"] = None

    for row in rows:
        key = canonical_client_key(row.get("client_name"))
        ids = keys.get(key, set())
        if key in unique:
            row["client_id"] = unique[key]
            matched += 1
        elif len(ids) > 1:
            ambiguous.add(row.get("client_name") or "")
        else:
            unresolved.add(row.get("client_name") or "")

    log(f"  Прямая привязка purchase_history.client_id: {matched} из {len(rows)} строк")
    if ambiguous:
        log(f"  ⚠️ Неоднозначные названия ({len(ambiguous)}) — client_id не проставлен:")
        for name in sorted(ambiguous)[:20]:
            log(f"    ! {name}")
    if unresolved:
        log(f"  ⚠️ Пока не сопоставлены с карточкой ({len(unresolved)}):")
        for name in sorted(unresolved)[:20]:
            log(f"    ? {name}")


def ensure_clients_exist(rows):
    """Заводит в CRM клиентов, которые покупают, но карточки не имеют."""
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }

    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/clients?select=id,name,manager_name,assortment",
        headers=headers, timeout=60,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"Не удалось получить клиентов: {resp.status_code} {resp.text}")
    existing = resp.json()
    # Включаем прежние названия объединённых торговых точек. Главный источник —
    # client_aliases; скрытый хвост assortment остаётся запасным вариантом.
    aliases = []
    try:
        aliases = _get_all_rest_rows("client_aliases", "client_id,alias_name")
    except Exception as exc:
        log(f"  ⚠️ Алиасы клиентов не прочитаны при проверке дублей: {exc}")

    aliases_by_id = {}
    for a in aliases:
        aliases_by_id.setdefault(str(a.get("client_id") or ""), []).append(a.get("alias_name") or "")

    existing_norm = []
    for c in existing:
        names = client_name_variants(c) + aliases_by_id.get(str(c.get("id") or ""), [])
        existing_norm.extend(normalize_name(n) for n in names if normalize_name(n))

    # Кто покупает по данным отчёта (имя -> менеджер + выручка за месяц).
    from_1c = {}
    for r in rows:
        key = r["client_name"]
        d = from_1c.setdefault(key, {"manager": r["manager_name"], "revenue": 0.0})
        d["revenue"] += r["revenue"]

    to_create = []
    seen = set()      # защита от дублей внутри самого отчёта
    fuzzy = []        # неточные совпадения — их стоит показать человеку

    for name, info in from_1c.items():
        norm = normalize_name(name)
        if not norm or norm in seen:
            continue

        exact = any(norm == e for e in existing_norm)
        if exact:
            continue

        # Совпадение по ВХОЖДЕНИЮ — рискованное место: "иванов" входит в
        # "иванова", и это разные ИП. Автоматически различить нельзя, поэтому
        # НЕ заводим (лучше пропустить, чем создать дубль), но сообщаем.
        near = [e for e in existing_norm if names_match(norm, e)]
        if near:
            fuzzy.append((name, near[0]))
            continue

        seen.add(norm)
        to_create.append({
            "name": name.strip(),
            "manager_name": info["manager"],
            "client_status": "Рабочий",   # он покупает — значит рабочий
            "auto_created": True,
            "created_from": "1С: продажи",
            "reviewed": False,
        })

    if fuzzy:
        log(f"  ⚠️ Похожи на существующих — НЕ заведены, проверьте вручную ({len(fuzzy)}):")
        for name, near in fuzzy:
            log(f"    ? «{name}» похож на «{near}»")

    if not to_create:
        log("  Новых клиентов из 1С нет — все уже заведены.")
        return

    log(f"  НОВЫЕ КЛИЕНТЫ ИЗ 1С: {len(to_create)}")
    for c in to_create:
        log(f"    + {c['name']} ({c['manager_name']}, закупки {from_1c[c['name']]['revenue']:,.2f} BYN)")

    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/clients",
        headers={**headers, "Prefer": "return=minimal"},
        json=to_create, timeout=60,
    )
    if resp.status_code not in (200, 201, 204):
        # Не валим весь импорт: продажи важнее. Клиентов заведём в следующий раз.
        log(f"  ⚠️ Не удалось завести клиентов: {resp.status_code} {resp.text}")
        return
    log(f"  Заведено карточек: {len(to_create)}. Разберите их в CRM "
        f"(адрес, категория, ассортимент) — раздел «Контроль».")


def _get_all_rest_rows(table, select):
    """Читает большую таблицу Supabase постранично (PostgREST Range)."""
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    out = []
    page = 1000
    start = 0
    while True:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/{table}?select={select}",
            headers={**headers, "Range": f"{start}-{start + page - 1}"},
            timeout=120,
        )
        if resp.status_code not in (200, 206):
            raise RuntimeError(f"Не удалось прочитать {table}: {resp.status_code} {resp.text}")
        chunk = resp.json() or []
        out.extend(chunk)
        if len(chunk) < page:
            break
        start += page
    return out


def refresh_merged_client_turnover():
    """Устаревшая функция оставлена для совместимости.

    ВАЖНО: ничего не записывает в clients.revenue_total. Источник правды —
    purchase_history, а CRM пересчитывает оборот после входа. Это защищает от
    массового обнуления, если REST-чтение истории вернуло пустой набор из-за RLS.
    """
    log("  Оборот клиентов пересчитывает CRM из purchase_history; поле clients.revenue_total не изменяется.")

def replace_month(month, rows):
    """Атомарно заменяет данные месяца через SQL-функцию Supabase.

    Удаление и вставка выполняются в ОДНОЙ транзакции. Если хотя бы одна строка
    некорректна, старая история месяца остаётся на месте. Это устраняет риск,
    при котором прежний импорт сначала удалял месяц, а затем мог упасть.
    """
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }

    if not rows:
        raise RuntimeError(
            f"Из отчёта не разобрано ни одной строки продаж за {month}. "
            f"Данные в CRM НЕ тронуты. Вероятно, в 1С изменилась структура отчёта."
        )

    # Единая схема всех JSON-объектов. Даже несопоставленная строка содержит
    # client_id=null — PostgREST больше не получает разные наборы ключей.
    fields = (
        "client_name", "category", "subgroup", "product", "sku", "month",
        "qty", "revenue", "manager_name", "client_id",
    )
    clean_rows = [{field: row.get(field) for field in fields} for row in rows]

    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/rpc/replace_purchase_history_month_safe",
        headers=headers,
        json={"p_month": month, "p_rows": clean_rows},
        timeout=180,
    )
    if resp.status_code not in (200, 201, 204):
        if resp.status_code in (404, 400) and (
            "PGRST202" in resp.text or "replace_purchase_history_month_safe" in resp.text
        ):
            raise RuntimeError(
                "В Supabase не установлена безопасная функция импорта. "
                "Сначала запустите install-safe-sales-import.sql, затем повторите workflow."
            )
        raise RuntimeError(
            f"Не удалось безопасно заменить продажи за {month}: "
            f"{resp.status_code} {resp.text}"
        )

    try:
        inserted = int(resp.json())
    except Exception:
        inserted = len(clean_rows)
    log(f"  Безопасно заменены данные за {month}: {inserted} строк")


# ===== v22.7.13: точный YoY ВИП на ту же дату =====
# purchase_history хранит только агрегат месяца и ежедневно ПЕРЕЗАПИСЫВАЕТ текущий
# месяц. Поэтому после закрытия августа 2025 из неё уже нельзя восстановить состояние
# на 07.08.2025. Отдельный снимок сохраняет только строки ВИП-клиентов и только на
# фактическую конечную дату отчёта — без экстраполяции и без изменения основной истории.

def _vip_snapshot_context():
    try:
        defs = _get_all_rest_rows("vip_sales", "client_name")
    except Exception as exc:
        log(f"  ⚠️ ВИП-снимок: список vip_sales недоступен: {exc}")
        return set(), []
    vip_names = [str(r.get("client_name") or "").strip() for r in defs if str(r.get("client_name") or "").strip()]
    if not vip_names:
        return set(), []

    vip_norm = [normalize_name(n) for n in vip_names if normalize_name(n)]
    vip_ids = set()
    try:
        clients = _get_all_rest_rows("clients", "id,name,assortment")
        aliases = []
        try:
            aliases = _get_all_rest_rows("client_aliases", "client_id,alias_name")
        except Exception:
            pass
        aliases_by_id = {}
        for a in aliases:
            aliases_by_id.setdefault(str(a.get("client_id") or ""), []).append(a.get("alias_name") or "")
        for c in clients:
            cid = str(c.get("id") or "")
            variants = client_name_variants(c) + aliases_by_id.get(cid, [])
            norms = [normalize_name(v) for v in variants if normalize_name(v)]
            if any(names_match(v, w) for v in norms for w in vip_norm):
                vip_ids.add(cid)
    except Exception as exc:
        log(f"  ⚠️ ВИП-снимок: client_id определяются только по именам: {exc}")
    return vip_ids, vip_norm


def _vip_rows_only(rows):
    vip_ids, vip_norm = _vip_snapshot_context()
    if not vip_norm:
        return []
    out = []
    for row in rows:
        cid = str(row.get("client_id") or "")
        rn = normalize_name(row.get("client_name"))
        if (cid and cid in vip_ids) or (rn and any(names_match(rn, v) for v in vip_norm)):
            out.append(row)
    return out


def vip_snapshot_exists(snapshot_date):
    try:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/vip_yoy_snapshot_runs",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
            params={"snapshot_date": f"eq.{snapshot_date.isoformat()}", "select": "snapshot_date", "limit": "1"},
            timeout=30,
        )
        if resp.status_code == 200:
            return bool(resp.json())
        return False
    except Exception:
        return False


def save_vip_yoy_snapshot(snapshot_date, month, rows, *, source_message_at=None):
    vip_rows = _vip_rows_only(rows)
    if not vip_rows:
        log("  ⚠️ ВИП-снимок не сохранён: среди строк отчёта не найдено ВИП-клиентов.")
        return False
    fields = (
        "client_name", "category", "subgroup", "product", "sku", "month",
        "qty", "revenue", "manager_name", "client_id",
    )
    clean = [{field: row.get(field) for field in fields} for row in vip_rows]
    try:
        resp = requests.post(
            f"{SUPABASE_URL}/rest/v1/rpc/replace_vip_yoy_snapshot_safe",
            headers=_api_headers(),
            json={
                "p_snapshot_date": snapshot_date.isoformat(),
                "p_month": month,
                "p_rows": clean,
                "p_source_message_at": source_message_at,
            },
            timeout=120,
        )
        if resp.status_code not in (200, 201, 204):
            log(f"  ⚠️ ВИП YoY-снимок не записан: {resp.status_code} {resp.text[:300]}")
            return False
        log(f"  ✅ ВИП YoY-снимок на {snapshot_date:%d.%m.%Y}: {len(clean)} SKU-строк")
        return True
    except Exception as exc:
        log(f"  ⚠️ ВИП YoY-снимок не записан: {exc}")
        return False


def backfill_previous_year_snapshot(mail, current_end):
    """Пытается автоматически найти в почте отчёт ровно на ту же дату год назад.

    Это разовая страховка для уже прошедшего 2025 года. Если такой ежедневной
    рассылки тогда не было, импорт продаж всё равно остаётся успешным, а CRM
    просто не показывает ложный процент до ручной загрузки исторического среза.
    """
    target_end = _previous_year_same_date(current_end)
    if vip_snapshot_exists(target_end):
        log(f"  Исторический ВИП-снимок {target_end:%d.%m.%Y} уже есть.")
        return True

    since = (target_end - timedelta(days=2)).strftime("%d-%b-%Y")
    before = (target_end + timedelta(days=4)).strftime("%d-%b-%Y")
    try:
        status, data = mail.search(None, f'(SINCE {since} BEFORE {before})')
        if status != "OK":
            return False
        candidates = []
        for msg_id in data[0].split():
            st, msg_data = mail.fetch(msg_id, "(RFC822)")
            if st != "OK":
                continue
            msg = email.message_from_bytes(msg_data[0][1])
            raw_subject = msg.get("Subject", "")
            subject = "".join(
                part.decode(enc or "utf-8", errors="replace") if isinstance(part, bytes) else part
                for part, enc in decode_header(raw_subject)
            )
            if SUBJECT_MARKER.lower() not in subject.lower():
                continue
            filename, content = find_xlsx_in_email(msg)
            if not content:
                continue
            try:
                start, end = parse_report_bounds(content)
            except Exception:
                continue
            if end != target_end or start != date(target_end.year, target_end.month, 1):
                continue
            sent = normalize_message_date(msg.get("Date"))
            candidates.append((sent, filename, content))
        if not candidates:
            log(f"  ⚠️ В почте не найден отчёт 1С за 01–{target_end:%d.%m.%Y}. Точный YoY пока будет без процента.")
            return False
        candidates.sort(key=lambda x: x[0])
        sent, filename, content = candidates[-1]
        month, hist_rows = parse_sales(content)
        try:
            attach_client_ids(hist_rows)
        except Exception as exc:
            log(f"  ⚠️ Исторический client_id: {exc}")
        ok = save_vip_yoy_snapshot(target_end, month, hist_rows, source_message_at=sent.isoformat())
        if ok:
            log(f"  Историческая база взята из письма «{filename}» от {sent:%d.%m.%Y %H:%M}.")
        return ok
    except Exception as exc:
        log(f"  ⚠️ Автопоиск исторического ВИП-среза пропущен: {exc}")
        return False

def find_xlsx_in_email(msg):
    for part in msg.walk():
        filename = part.get_filename()
        if not filename:
            continue
        decoded = decode_header(filename)[0]
        if isinstance(decoded[0], bytes):
            filename = decoded[0].decode(decoded[1] or "utf-8", errors="replace")
        if filename.lower().endswith((".xlsx", ".xls")):
            return filename, part.get_payload(decode=True)
    return None, None


def check_period(month):
    """Сверяет месяц отчёта с текущим.

    Зачем: у отчёта продаж в 1С есть параметр ПЕРИОДА. Если он зафиксирован
    конкретными датами (01.07.2026 - 31.07.2026), то с наступлением августа
    рассылка продолжит исправно приходить — но с июльскими данными. Продажи
    в CRM просто застынут, и никто этого не заметит. Ровно та же болезнь,
    что была у ПДЗ с «Произвольной датой».

    Допускаем текущий месяц и предыдущий: в первых числах 1С может присылать
    закрытие прошлого месяца, и это нормально.
    """
    today = date.today()
    cur = date(today.year, today.month, 1)
    prev = date(cur.year - 1, 12, 1) if cur.month == 1 else date(cur.year, cur.month - 1, 1)

    rep = datetime.strptime(month, "%Y-%m-%d").date()

    if rep == cur:
        log(f"  Период отчёта: {rep:%m.%Y} — текущий месяц, всё верно.")
        return 'current'
    if rep == prev:
        log(f"  ⚠️ Период отчёта: {rep:%m.%Y} — это ПРОШЛЫЙ месяц. Гружу (закрытие месяца), "
            f"но если так придёт и завтра — проверьте период в рассылке 1С.")
        return 'previous'

    raise RuntimeError(
        f"Отчёт за {rep:%m.%Y}, а сейчас {cur:%m.%Y}. Данные НЕ загружены, чтобы не "
        f"перезаписать историю устаревшими цифрами. Причина обычно одна: в 1С у отчёта "
        f"«Продажи для CRM» период задан конкретными датами вместо «Этот месяц» — "
        f"проверьте параметр периода В ОТЧЁТЕ И В РАССЫЛКЕ."
    )


def main():
    safe_set_import_status("sales", "running")
    log(f"Подключаюсь к почте {IMAP_HOST}:{IMAP_PORT} (таймаут {MAIL_TIMEOUT} сек.)")
    mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT, timeout=MAIL_TIMEOUT)
    mail.login(IMAP_USER, IMAP_PASS)
    select_mailbox(mail)
    try:
        mail.sock.settimeout(MAIL_TIMEOUT)
    except Exception:
        pass

    since = (date.today() - timedelta(days=LOOKBACK_DAYS)).strftime("%d-%b-%Y")
    status, data = mail.search(None, f'(SINCE {since})')
    if status != "OK":
        raise RuntimeError("Не удалось получить список писем")

    ids = data[0].split()
    log(f"Писем за последние {LOOKBACK_DAYS} дн.: {len(ids)}. Заголовки получаю пакетами по {HEADER_BATCH_SIZE}, без сотен отдельных IMAP-запросов.")
    candidates = sales_mail_candidates(mail, ids)

    # Полное письмо скачиваем начиная с самого свежего. Если первый отчёт
    # корректен и относится к текущему месяцу — дальше Gmail вообще не трогаем.
    # В начале месяца допустим свежий отчёт закрытия прошлого месяца: его держим
    # как fallback и продолжаем искать текущий среди следующих кандидатов.
    parsed_current = None
    parsed_previous = None
    parse_errors = []
    today = date.today()
    cur_month = f"{today.year:04d}-{today.month:02d}-01"
    prev_date = date(today.year - 1, 12, 1) if today.month == 1 else date(today.year, today.month - 1, 1)
    prev_month = prev_date.isoformat()

    for idx, meta in enumerate(candidates, 1):
        msg_id, sent, subject = meta["id"], meta["sent"], meta["subject"]
        log(f"  Проверяю полное письмо {idx}/{len(candidates)}: {sent:%d.%m.%Y %H:%M} — {subject}")
        try:
            status, msg_data = mail.fetch(msg_id, "(RFC822)")
        except Exception as exc:
            parse_errors.append(f"{sent:%d.%m %H:%M}: IMAP {exc}")
            continue
        if status != "OK" or not msg_data:
            parse_errors.append(f"{sent:%d.%m %H:%M}: Gmail не вернул полное письмо")
            continue
        raw = next((item[1] for item in msg_data if isinstance(item, tuple) and len(item)>1 and isinstance(item[1], (bytes, bytearray))), None)
        if not raw:
            parse_errors.append(f"{sent:%d.%m %H:%M}: пустое тело письма")
            continue
        msg = email.message_from_bytes(raw)
        filename, content = find_xlsx_in_email(msg)
        if not content:
            parse_errors.append(f"{sent:%d.%m %H:%M}: нет Excel-вложения")
            continue
        try:
            month, rows = parse_sales(content)
            _period_start, period_end = parse_report_bounds(content)
            parsed = (month, period_end, sent, subject, filename, rows)
        except Exception as exc:
            parse_errors.append(f"{sent:%d.%m %H:%M}: {exc}")
            continue

        if month == cur_month:
            parsed_current = parsed
            log("  ✅ Найден корректный отчёт текущего месяца — дальнейший поиск писем остановлен.")
            break
        if month == prev_month and parsed_previous is None:
            parsed_previous = parsed
            log("  ℹ️ Это закрытие прошлого месяца; сохраняю как запасной вариант и ищу текущий месяц.")
            continue
        parse_errors.append(f"{sent:%d.%m %H:%M}: период {month[:7]} не текущий и не предыдущий")

    selected = parsed_current or parsed_previous
    if not selected:
        detail = "; ".join(parse_errors[-3:])
        raise RuntimeError(
            f"Не найден пригодный отчёт «{SUBJECT_MARKER}» за последние {LOOKBACK_DAYS} дн."
            + (f" Последние ошибки: {detail}" if detail else "")
        )

    month, report_end, sent, subject, filename, rows = selected
    IMPORT_CONTEXT.update({
        "report_period": month[:7],
        "report_date": report_end.isoformat(),
        "source_message_at": sent.isoformat(),
        "row_count": len(rows),
    })
    log(f"Выбран отчёт: «{subject}» от {sent:%d.%m.%Y %H:%M}")
    log(f"  Период: {month[:7]} по {report_end:%d.%m.%Y}, вложение: {filename}, строк: {len(rows)}")

    period_state = check_period(month)

    # v23.6.60: почасовой workflow может видеть одно и то же письмо много раз.
    # Если оно уже успешно загружено, выходим без единой записи в БД и без
    # ложного realtime-сигнала для открытых экранов CRM.
    if sales_message_already_loaded(month[:7], sent):
        log(f"✅ Письмо {sent:%d.%m.%Y %H:%M} уже загружено — почасовая проверка без изменений.")
        mail.logout()
        return

    by_mgr = {}
    for row in rows:
        by_mgr.setdefault(row["manager_name"], {"rev": 0, "clients": set()})
        by_mgr[row["manager_name"]]["rev"] += row["revenue"]
        by_mgr[row["manager_name"]]["clients"].add(row["client_name"])
    for mgr, values in by_mgr.items():
        log(f"    {mgr}: {len(values['clients'])} клиентов, {values['rev']:,.2f} BYN")
    total_revenue = sum(row["revenue"] for row in rows)
    log(f"  ИТОГО выручка за {month}: {total_revenue:,.2f} BYN")

    try:
        ensure_clients_exist(rows)
    except Exception as exc:
        log(f"  ⚠️ Автозаведение клиентов не отработало: {exc}")

    try:
        attach_client_ids(rows)
    except Exception as exc:
        log(f"  ⚠️ Прямая привязка client_id не отработала: {exc}")

    replace_month(month, rows)

    # Не влияет на purchase_history: это отдельный точный MTD-снимок только для ВИП YoY.
    # Сохраняем текущий служебный ВИП-снимок, но больше НЕ ищем письмо
    # за ту же дату прошлого года: ВИП теперь сравнивается с полным месяцем 2025
    # из purchase_history, поэтому старый исторический IMAP-поиск только тратил время.
    save_vip_yoy_snapshot(report_end, month, rows, source_message_at=sent.isoformat())

    try:
        resp = requests.post(
            f"{SUPABASE_URL}/rest/v1/rpc/refresh_client_revenue_totals",
            headers=_api_headers(), json={}, timeout=120,
        )
        if resp.status_code not in (200, 201, 204):
            log(f"  ⚠️ Пересчёт оборотов после импорта пропущен: {resp.status_code} {resp.text}")
    except Exception as exc:
        log(f"  ⚠️ Пересчёт оборотов после импорта не выполнен: {exc}")

    capture_promotion_sales_snapshots(month, sent.isoformat())

    safe_set_import_status(
        "sales", "ok",
        report_period=month[:7],
        report_date=report_end.isoformat(),
        source_message_at=sent.isoformat(),
        row_count=len(rows),
        details=(f"Выручка {total_revenue:.2f} BYN; файл {filename}; "
                 + ("текущий месяц" if period_state=="current" else "закрытие прошлого месяца, ожидается текущий")),
    )
    mail.logout()
    log("Готово.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        safe_set_import_status(
            "sales", "error",
            report_period=IMPORT_CONTEXT.get("report_period"),
            report_date=IMPORT_CONTEXT.get("report_date"),
            source_message_at=IMPORT_CONTEXT.get("source_message_at"),
            row_count=IMPORT_CONTEXT.get("row_count"),
            error_text=str(e)[:2000],
        )
        log(f"ОШИБКА: {e}")
        sys.exit(1)
