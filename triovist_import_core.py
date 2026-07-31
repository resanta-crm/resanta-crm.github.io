#!/usr/bin/env python3
"""Общий безопасный парсер отчётов 1С для ООО «Триовист».

Деньги читаются через Decimal и отправляются строками, чтобы сохранить копейки.
Распределение между менеджерами НЕ берётся из колонок старых менеджеров в 1С:
его выполняет Supabase по действующей таблице triovist_group_assignments.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable

import openpyxl

ARTICLE_RE = re.compile(r"^\d+(?:/\d+){1,6}$")
TRIOVIST_NAMES = {"ооотриовист", "триовистооо", "триовист"}
MONTHS_RU = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}
SKIP_NAMES = {"Контрагент", "Номенклатура", "Параметры:", "Отбор:", "Итого", "Артикул"}


def log(message: str) -> None:
    print(message, flush=True)


def norm_name(value: object) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"[^0-9a-zа-я]+", "", text)
    return text


def is_triovist(value: object) -> bool:
    return norm_name(value) in TRIOVIST_NAMES


def dec(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise RuntimeError(f"Не удалось прочитать число: {value!r}") from exc


def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def quantity(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)


def month_key(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}-01"


def parse_month_header(value: object) -> str | None:
    text = str(value or "").strip().lower().replace("ё", "е")
    match = re.match(
        r"^(январь|февраль|март|апрель|май|июнь|июль|август|сентябрь|октябрь|ноябрь|декабрь)\s+(\d{4})",
        text,
    )
    if not match:
        return None
    return month_key(int(match.group(2)), MONTHS_RU[match.group(1)])


def parse_report_period(ws) -> str:
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=min(ws.max_column, 12)):
        for cell in row:
            if not cell.value:
                continue
            match = re.search(r"Период:\s*(\d{2})\.(\d{2})\.(\d{4})", str(cell.value))
            if match:
                _day, month, year = match.groups()
                return f"{year}-{month}-01"
    raise RuntimeError("Не найден период отчёта 1С")


def _is_qty(value: object) -> bool:
    return str(value or "").strip().lower().startswith("кол")


def _is_revenue(value: object) -> bool:
    return str(value or "").strip().lower().startswith("выруч")


def find_month_columns(ws, target_month: str | None = None) -> list[tuple[str, int, int]]:
    result: list[tuple[str, int, int]] = []
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        for col_idx in range(1, ws.max_column):
            parsed = parse_month_header(ws.cell(row_idx, col_idx).value)
            if not parsed or (target_month and parsed != target_month):
                continue
            if _is_qty(ws.cell(row_idx + 1, col_idx).value) and _is_revenue(ws.cell(row_idx + 1, col_idx + 1).value):
                result.append((parsed, col_idx, col_idx + 1))
    # Убираем возможные дубли, сохраняя порядок.
    unique: dict[str, tuple[str, int, int]] = {}
    for item in result:
        unique[item[0]] = item
    return list(unique.values())


@dataclass(frozen=True)
class BodyRow:
    row: int
    name: str
    indent: int


def build_body(ws) -> list[BodyRow]:
    body: list[BodyRow] = []
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value is None or not str(cell.value).strip():
            continue
        name = str(cell.value).strip()
        if name in SKIP_NAMES:
            continue
        body.append(BodyRow(row_idx, name, int(cell.alignment.indent or 0)))
    return body


def parse_rows_for_month(ws, month: str, qty_col: int, revenue_col: int) -> tuple[list[dict], Decimal, Decimal]:
    body = build_body(ws)
    if not body:
        raise RuntimeError("Отчёт пустой")

    product_indexes = {
        idx for idx in range(len(body) - 1)
        if not ARTICLE_RE.match(body[idx].name) and ARTICLE_RE.match(body[idx + 1].name)
    }
    client_indent = min(row.indent for row in body)
    client: str | None = None
    triovist_total_row: int | None = None
    heading_stack: dict[int, str] = {}
    rows: list[dict] = []

    for idx, item in enumerate(body):
        name = item.name
        if ARTICLE_RE.match(name):
            continue

        if idx in product_indexes:
            if not client or not is_triovist(client):
                continue
            headings = [heading_stack[key] for key in sorted(heading_stack) if key > client_indent]
            category = headings[0] if headings else "Без категории"
            subgroup = headings[-1] if len(headings) > 1 else (headings[0] if headings else "Прочее")
            sku = body[idx + 1].name if idx + 1 < len(body) and ARTICLE_RE.match(body[idx + 1].name) else ""
            qty = quantity(dec(ws.cell(item.row, qty_col).value))
            revenue = money(dec(ws.cell(item.row, revenue_col).value))
            if qty == 0 and revenue == 0:
                continue
            rows.append({
                "category": category,
                "subgroup": subgroup,
                "product": name,
                "sku": sku,
                "qty": format(qty, "f"),
                "revenue": format(revenue, "f"),
            })
            continue

        if item.indent <= client_indent:
            client = name
            heading_stack = {}
            if is_triovist(client):
                triovist_total_row = item.row
        else:
            for key in [key for key in heading_stack if key >= item.indent]:
                del heading_stack[key]
            heading_stack[item.indent] = name

    if triovist_total_row is None:
        raise RuntimeError("В отчёте не найден клиент «Триовист ООО»")
    if not rows:
        raise RuntimeError(f"У Триовиста не разобрано ни одной строки за {month}; база не изменена")

    expected_qty = quantity(dec(ws.cell(triovist_total_row, qty_col).value))
    expected_revenue = money(dec(ws.cell(triovist_total_row, revenue_col).value))
    parsed_qty = quantity(sum((dec(row["qty"]) for row in rows), Decimal("0")))
    parsed_revenue = money(sum((dec(row["revenue"]) for row in rows), Decimal("0")))

    if parsed_qty != expected_qty or parsed_revenue != expected_revenue:
        raise RuntimeError(
            f"Контроль отчёта не сошёлся за {month}: "
            f"1С={expected_qty} шт. / {expected_revenue} BYN, "
            f"разобрано={parsed_qty} шт. / {parsed_revenue} BYN. База не изменена."
        )
    return rows, expected_qty, expected_revenue


def load_workbook_from_bytes(content: bytes):
    if content[:2] != b"PK":
        raise RuntimeError("Нужен отчёт .xlsx (Excel 2007 и новее), старый .xls не поддерживается")
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)


def load_workbook_from_path(path: str | Path):
    return openpyxl.load_workbook(Path(path), data_only=True, read_only=False)


def parse_current_report(content: bytes) -> tuple[str, list[dict], Decimal, Decimal]:
    wb = load_workbook_from_bytes(content)
    ws = wb.active
    month = parse_report_period(ws)
    columns = find_month_columns(ws, month)
    if len(columns) != 1:
        raise RuntimeError(f"Не найдена единственная пара Количество/Выручка за {month}: найдено {len(columns)}")
    _, qty_col, revenue_col = columns[0]
    rows, qty, revenue = parse_rows_for_month(ws, month, qty_col, revenue_col)
    return month, rows, qty, revenue


def parse_history_report(path: str | Path) -> list[tuple[str, list[dict], Decimal, Decimal]]:
    wb = load_workbook_from_path(path)
    ws = wb.active
    columns = sorted(find_month_columns(ws), key=lambda item: item[0])
    if not columns:
        raise RuntimeError("В историческом отчёте не найдены месяцы")
    result = []
    for month, qty_col, revenue_col in columns:
        rows, qty, revenue = parse_rows_for_month(ws, month, qty_col, revenue_col)
        result.append((month, rows, qty, revenue))
    return result


def check_current_or_previous(month: str) -> None:
    report = datetime.strptime(month, "%Y-%m-%d").date()
    today = date.today()
    current = date(today.year, today.month, 1)
    previous = date(current.year - 1, 12, 1) if current.month == 1 else date(current.year, current.month - 1, 1)
    if report not in (current, previous):
        raise RuntimeError(
            f"Рассылка прислала устаревший период {month[:7]}; ожидается {current:%Y-%m} "
            f"или закрытие {previous:%Y-%m}. Данные не изменены."
        )
