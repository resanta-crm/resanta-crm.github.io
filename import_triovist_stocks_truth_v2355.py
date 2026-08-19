#!/usr/bin/env python3
"""Triovist stock import truth v23.5.5.

Corrects the 21vek source semantics before staging:
- "Продажи за 3 мес." is a three-month total, not a third monthly value;
- "Статистика продаж1/2" are diagnostics only and never enter sales_90;
- 21vek available = free now + free in transit;
- total in transit is diagnostic only and does not cover demand unless it is free;
- raw source diagnostics and partner forecast are preserved in match_note;
- Chekhov parser/upload staging reuse the already proven atomic base implementation.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

from import_triovist_stocks import (
    dec,
    extract_internal_sku,
    is_group900_sku,
    log,
    norm,
    parse_chekhov,
    upload_staged,
)

VERSION = "v23.5.5"
SOURCE_MARK = "[truth-v23.5.5]"
SKU_RE = re.compile(r"^\d+(?:/\d+){1,6}$")


def _n(value) -> float:
    return float(dec(value))


def _header_index(headers: list[str], *names: str, starts: bool = True) -> int:
    normalized = [norm(v) for v in headers]
    for name in names:
        key = norm(name)
        for idx, value in enumerate(normalized):
            if value == key or (starts and value.startswith(key)):
                return idx
    return -1


def _extract_sku(value) -> str:
    text = str(value or "").strip().lstrip("'")
    return text if SKU_RE.fullmatch(text) else ""


def _find_header(ws) -> tuple[int, list[str]]:
    for row_idx, values in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 35), values_only=True), start=1):
        headers = [str(v or "").strip() for v in values]
        keys = [norm(v) for v in headers]
        has_product = any(v == "номенклатура" or v.startswith("номенклатура") or v.startswith("наименование") for v in keys)
        has_total = any(v == "всего" or v.startswith("всего") for v in keys)
        if has_product and has_total:
            return row_idx, headers
    raise RuntimeError("Не найдена шапка 21vek")


def parse_partner_truth(path: str | Path):
    wb = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    ws = wb.active
    header_row, headers = _find_header(ws)
    keys = [norm(v) for v in headers]

    c_product = _header_index(headers, "Номенклатура", "Наименование")
    c_partner = _header_index(headers, "Артикул", "Артикул 21vek", "Код товара", "ID товара", "SKU")
    c_kind = _header_index(headers, "Вид", "Категория")
    c_brand = _header_index(headers, "Производитель", "Бренд")
    c_sales3 = _header_index(headers, "Продажи за 3 мес.", "Продажи за 3 месяца", "Продажи 3 мес.")
    legacy_sales = [i for i, value in enumerate(keys) if value.startswith("продажиза") and not value.startswith("продажиза3мес")][-3:]
    if c_sales3 < 0 and len(legacy_sales) < 3:
        raise RuntimeError(
            "Не найдены продажи 21vek: нужна колонка «Продажи за 3 мес.» либо три старые колонки «Продажи за ...». "
            "«Статистика продаж» намеренно не используется."
        )
    c_stat2 = _header_index(headers, "Статистика продаж2")
    c_stat1 = _header_index(headers, "Статистика продаж1")
    c_total = _header_index(headers, "Всего", "Остаток всего")
    c_reserve = _header_index(headers, "Резерв")
    c_transit = _header_index(headers, "В пути")
    c_orders = _header_index(headers, "Заказ", "Заказы", "В заказах")
    c_forecast = _header_index(headers, "Прогноз")

    free_exact = [i for i, value in enumerate(keys) if value == "свободно"]
    c_free_transit = _header_index(headers, "Свободно в пути", "Свободно в пути и резерве")
    c_free_now = -1
    if c_transit >= 0:
        before = [i for i in free_exact if i < c_transit]
        after = [i for i in free_exact if i > c_transit]
        c_free_now = before[-1] if before else -1
        if c_free_transit < 0 and after:
            c_free_transit = after[0]
    elif free_exact:
        c_free_now = free_exact[0]
    if c_free_now < 0:
        c_free_now = _header_index(headers, "Доступно")

    explicit_cols = []
    for name in (
        "Наш артикул", "Артикул поставщика", "Артикул производителя", "SKU поставщика",
        "SKU производителя", "Код поставщика", "Код производителя", "Наш SKU",
    ):
        idx = _header_index(headers, name)
        if idx >= 0 and idx not in explicit_cols:
            explicit_cols.append(idx)

    rows = []
    diagnostics = {
        "version": VERSION,
        "header_row": header_row,
        "sales_mode": "three_month_total" if c_sales3 >= 0 else "legacy_three_months",
        "sales_column": headers[c_sales3] if c_sales3 >= 0 else " + ".join(headers[i] for i in legacy_sales),
        "stat1_column": headers[c_stat1] if c_stat1 >= 0 else None,
        "stat2_column": headers[c_stat2] if c_stat2 >= 0 else None,
        "forecast_column": headers[c_forecast] if c_forecast >= 0 else None,
        "free_now_column": headers[c_free_now] if c_free_now >= 0 else None,
        "free_transit_column": headers[c_free_transit] if c_free_transit >= 0 else None,
        "samples": [],
    }

    for row_no, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = list(values)
        product = str(values[c_product] or "").strip() if c_product >= 0 and c_product < len(values) else ""
        if not product:
            continue
        kind = str(values[c_kind] or "").strip() if c_kind >= 0 and c_kind < len(values) else ""
        packed = norm(kind + " " + product)
        excluded = "+" in product or "комплект" in packed or "невыводить" in norm(product)

        sku = ""
        for idx in explicit_cols:
            if idx < len(values):
                sku = _extract_sku(values[idx])
                if sku:
                    break
        if not sku and c_partner >= 0 and c_partner < len(values):
            sku = _extract_sku(values[c_partner])
        if not sku:
            sku = extract_internal_sku(product)
        if not sku:
            if excluded:
                continue
            continue
        if is_group900_sku(sku):
            continue

        raw_total = max(0.0, _n(values[c_total]) if c_total >= 0 and c_total < len(values) else 0.0)
        reserve = max(0.0, _n(values[c_reserve]) if c_reserve >= 0 and c_reserve < len(values) else 0.0)
        free_now = max(0.0, _n(values[c_free_now]) if c_free_now >= 0 and c_free_now < len(values) else raw_total - reserve)
        transit = max(0.0, _n(values[c_transit]) if c_transit >= 0 and c_transit < len(values) else 0.0)
        free_transit_raw = max(0.0, _n(values[c_free_transit]) if c_free_transit >= 0 and c_free_transit < len(values) else 0.0)
        # Physical invariant: free stock in transit can never exceed total stock in transit.
        # Some 21vek files reuse a second column named «Свободно» for another balance field;
        # if «В пути» is empty/zero we must not count that second value as transit.
        free_transit = min(free_transit_raw, transit) if c_transit >= 0 else free_transit_raw
        available = free_now + free_transit
        sales_90 = max(0.0, _n(values[c_sales3])) if c_sales3 >= 0 else max(0.0, sum(_n(values[i]) for i in legacy_sales))
        orders = max(0.0, _n(values[c_orders]) if c_orders >= 0 and c_orders < len(values) else 0.0)
        partner_forecast = _n(values[c_forecast]) if c_forecast >= 0 and c_forecast < len(values) else None
        partner_sku = str(values[c_partner] or "").strip() if c_partner >= 0 and c_partner < len(values) else sku

        note = {
            "truth_version": VERSION,
            "sales_3m": sales_90,
            "stat_sales_2": _n(values[c_stat2]) if c_stat2 >= 0 and c_stat2 < len(values) else None,
            "stat_sales_1": _n(values[c_stat1]) if c_stat1 >= 0 and c_stat1 < len(values) else None,
            "total": raw_total,
            "free_now": free_now,
            "reserve": reserve,
            "in_transit": transit,
            "free_in_transit_raw": free_transit_raw,
            "free_in_transit": free_transit,
            "partner_forecast": partner_forecast,
        }
        rows.append({
            "row_key": f"{partner_sku or sku}|{sku}|{product}",
            "partner_sku": partner_sku or sku,
            "sku": sku,
            "kind": kind,
            "brand": str(values[c_brand] or "").strip() if c_brand >= 0 and c_brand < len(values) else "",
            "product": product,
            "sales_m1": "0",
            "sales_m2": "0",
            "sales_m3": str(sales_90),
            "sales_m1_label": "CRM truth: не используется",
            "sales_m2_label": "CRM truth: не используется",
            "sales_m3_label": headers[c_sales3] if c_sales3 >= 0 else "CRM truth: сумма 3 старых месяцев",
            "qty_total": str(available),
            "qty_free": str(available),
            "qty_transit_reserve": str(free_transit + reserve),
            "qty_orders": str(orders),
            "excluded": excluded,
            "match_note": json.dumps(note, ensure_ascii=False, separators=(",", ":")),
        })
        if len(diagnostics["samples"]) < 8 or sku in {"70/1/65", "72/17/1", "72/11/6", "65/60"}:
            diagnostics["samples"].append({"row": row_no, "sku": sku, **note, "available": available, "orders": orders})
            diagnostics["samples"] = diagnostics["samples"][-12:]

    if len(rows) < 100:
        raise RuntimeError(f"Слишком мало корректных SKU 21vek: {len(rows)}")
    return rows, diagnostics


def main():
    partner = Path(sys.argv[1] if len(sys.argv) > 1 else "data/stock_21vek.xlsx")
    chekhov = Path(sys.argv[2] if len(sys.argv) > 2 else "data/stock_chekhov.xlsx")
    partner_rows, diag = parse_partner_truth(partner)
    chekhov_rows = parse_chekhov(chekhov)
    log("21vek truth: " + json.dumps(diag, ensure_ascii=False))
    log(f"21vek: {len(partner_rows)} строк; Чехов: {len(chekhov_rows)} строк")
    log(f"21vek RPC: {upload_staged('partner', partner_rows, SOURCE_MARK + ' ' + partner.name)}")
    log(f"Чехов RPC: {upload_staged('chekhov', chekhov_rows, chekhov.name)}")
    log(f"✅ Остатки загружены по правилам {VERSION}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log(f"ОШИБКА: {exc}")
        raise
