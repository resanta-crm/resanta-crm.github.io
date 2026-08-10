#!/usr/bin/env python3
"""Import recent 1C shipments to ООО «Триовист» from Gmail into triovist_shipments.

v22.7.30.1 — Gmail report discovery fixed:
- Gmail endpoint remains pinned to imap.gmail.com:993 with bounded retries;
- Gmail X-GM-RAW first selects only recent Excel attachments and explicitly excludes GitHub notifications;
- Subject is only a ranking hint, never a hard requirement: the Excel structure decides whether this is a 1C shipment report;
- every Excel attachment in candidate emails is checked, not only the first one;
- all valid recent reports are merged by document/date/SKU, newest report wins on duplicates;
- no valid report is a normal green run and never changes existing shipment data.

Expected Excel: at minimum Артикул + Количество. Optional columns:
Дата / Дата документа, Документ / Номер документа, Номенклатура / Товар,
Контрагент / Клиент, Сумма / Выручка.
If Контрагент exists, only rows containing "Триовист" are imported.
Group/SKU 900 is always ignored.
"""
from __future__ import annotations

import email
import imaplib
import io
import os
import re
import socket
import sys
import time
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from email.header import decode_header
from email.utils import parsedate_to_datetime

import openpyxl
import requests

VERSION = "v22.7.30.1"
GMAIL_HOST = "imap.gmail.com"
GMAIL_PORT = 993


def _env_text(name: str, default: str = "") -> str:
    value = os.getenv(name)
    if value is None:
        return default
    value = str(value).strip()
    return value or default


def _safe_port(value: str | None, default: int = GMAIL_PORT) -> int:
    try:
        port = int(str(value or "").strip())
        if 1 <= port <= 65535:
            return port
    except Exception:
        pass
    return default


# This importer is Gmail-specific.  An empty GitHub secret used to overwrite the
# Python default and caused [Errno 111] Connection refused.  Keep a defensive
# fallback here as well as a fixed endpoint in the workflow.
IMAP_HOST = _env_text("IMAP_HOST", GMAIL_HOST)
IMAP_PORT = _safe_port(os.getenv("IMAP_PORT"), GMAIL_PORT)
if IMAP_HOST.lower() in {"localhost", "127.0.0.1", "0.0.0.0", "none", "null", "-"}:
    IMAP_HOST = GMAIL_HOST
if IMAP_PORT != GMAIL_PORT:
    # Gmail IMAP SSL is expected on 993.  Do not let an accidental shared secret
    # point this dedicated importer at another service/port.
    IMAP_PORT = GMAIL_PORT

IMAP_USER = os.environ["IMAP_USER"].strip()
IMAP_PASS = "".join(os.environ["IMAP_PASS"].split()).replace("\xa0", "")
SUPABASE_URL = os.environ["SUPABASE_URL"].strip().rstrip("/")
SUPABASE_KEY = os.environ["SUPABASE_KEY"].strip()
LOOKBACK = int(_env_text("TRIOVIST_SHIPMENTS_LOOKBACK_DAYS", "14"))
SUBJECT_KEYS = tuple(
    x.strip().lower()
    for x in _env_text("TRIOVIST_SHIPMENTS_SUBJECT_KEYS", "отгруз,триовист").split(",")
    if x.strip()
)
MAIL_TIMEOUT = max(15, int(_env_text("TRIOVIST_SHIPMENTS_IMAP_TIMEOUT_SECONDS", "45")))
CONNECT_ATTEMPTS = max(1, min(5, int(_env_text("TRIOVIST_SHIPMENTS_IMAP_ATTEMPTS", "3"))))
CONNECT_RETRY_SECONDS = max(1, min(30, int(_env_text("TRIOVIST_SHIPMENTS_IMAP_RETRY_SECONDS", "6"))))
HEADER_BATCH_SIZE = max(20, int(_env_text("TRIOVIST_SHIPMENTS_IMAP_HEADER_BATCH", "75")))
HEADER_SCAN_DEADLINE = max(30, int(_env_text("TRIOVIST_SHIPMENTS_HEADER_DEADLINE_SECONDS", "150")))
MAX_FULL_MESSAGES = max(1, min(60, int(_env_text("TRIOVIST_SHIPMENTS_MAX_EMAIL_CANDIDATES", "30"))))
ARTICLE_RE = re.compile(r"^\d+(?:/\d+){1,6}$")


def log(x):
    print(x, flush=True)


def norm(v):
    return re.sub(r"\s+", " ", str(v or "").strip().lower().replace("ё", "е"))


def dec(v):
    if v in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(v).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except InvalidOperation:
        return Decimal("0")


def decode(v):
    if not v:
        return ""
    out = []
    for part, enc in decode_header(v):
        if not isinstance(part, bytes):
            out.append(part)
            continue
        encoding = enc or "utf-8"
        if str(encoding).lower() == "unknown-8bit":
            encoding = "utf-8"
        try:
            out.append(part.decode(encoding, "replace"))
        except (LookupError, UnicodeDecodeError):
            out.append(part.decode("utf-8", "replace"))
    return "".join(out)


def msg_date(v):
    try:
        d = parsedate_to_datetime(v)
        return (d.replace(tzinfo=timezone.utc) if d.tzinfo is None else d).astimezone(timezone.utc)
    except Exception:
        return datetime.min.replace(tzinfo=timezone.utc)


def connect_mail() -> imaplib.IMAP4_SSL:
    """Connect/login with bounded retries for transient runner/Gmail failures."""
    log(f"{VERSION}: подключаюсь к почте {IMAP_HOST}:{IMAP_PORT}")
    last_exc: Exception | None = None
    for attempt in range(1, CONNECT_ATTEMPTS + 1):
        mail = None
        try:
            log(f"  IMAP попытка {attempt}/{CONNECT_ATTEMPTS}…")
            mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT, timeout=MAIL_TIMEOUT)
            mail.login(IMAP_USER, IMAP_PASS)
            try:
                mail.sock.settimeout(MAIL_TIMEOUT)
            except Exception:
                pass
            log("  ✅ Gmail IMAP подключён.")
            return mail
        except imaplib.IMAP4.error as exc:
            # Authentication/configuration errors are deterministic; retrying only
            # wastes the job and can trigger Gmail security throttling.
            try:
                if mail is not None:
                    mail.logout()
            except Exception:
                pass
            text = str(exc)
            if re.search(r"auth|authentication|credentials|login|invalid", text, re.I):
                raise RuntimeError(f"Gmail отклонил вход IMAP: {text}") from exc
            last_exc = exc
        except (imaplib.IMAP4.abort, socket.timeout, TimeoutError, ConnectionError, OSError) as exc:
            last_exc = exc
            try:
                if mail is not None:
                    mail.logout()
            except Exception:
                pass
        except Exception as exc:
            last_exc = exc
            try:
                if mail is not None:
                    mail.logout()
            except Exception:
                pass

        if attempt < CONNECT_ATTEMPTS:
            log(f"  ⚠️ IMAP временно недоступен: {last_exc}. Повтор через {CONNECT_RETRY_SECONDS} сек.")
            time.sleep(CONNECT_RETRY_SECONDS)

    raise RuntimeError(
        f"Не удалось подключиться к Gmail {IMAP_HOST}:{IMAP_PORT} после {CONNECT_ATTEMPTS} попыток: {last_exc}. "
        "Существующие отгрузки в CRM не изменялись."
    )


def select_box(mail):
    """Select Gmail All Mail; fall back to INBOX exactly like proven sales importers."""
    try:
        st, boxes = mail.list()
        if st == "OK":
            for raw in boxes or []:
                if b"\\All" not in raw:
                    continue
                m = re.search(br'\)\s+"[^"]*"\s+(.+)$', raw)
                if not m:
                    continue
                token = m.group(1).strip()
                if token.startswith(b'"') and token.endswith(b'"'):
                    token = token[1:-1]
                mailbox = token.decode("ascii", "ignore")
                if mailbox:
                    arg = f'"{mailbox}"' if " " in mailbox else mailbox
                    if mail.select(arg)[0] == "OK":
                        log("Ищу отгрузки во всей почте: " + mailbox)
                        return mailbox
    except Exception as e:
        log("  ⚠️ Не удалось открыть All Mail: " + str(e))
    if mail.select("INBOX")[0] != "OK":
        raise RuntimeError("Не удалось открыть Gmail All Mail или INBOX")
    log("Ищу отгрузки в INBOX")
    return "INBOX"


def attachments(msg):
    """Return all OpenXML Excel attachments; ignore inline/system MIME parts."""
    out = []
    for p in msg.walk():
        fn = decode(p.get_filename()).strip()
        if not fn or not fn.lower().endswith((".xlsx", ".xlsm")):
            continue
        payload = p.get_payload(decode=True)
        if payload:
            out.append((fn, payload))
    return out


SYSTEM_SENDER_RE = re.compile(r"(?:^|[<\s])(?:notifications|noreply)@github\.com(?:[>\s]|$)|github-actions", re.I)
SYSTEM_SUBJECT_RE = re.compile(r"(?:\[resanta-crm/|run failed:|run cancelled:|workflow run|github actions|action required)", re.I)


def is_system_notice(subject: str, sender: str) -> bool:
    return bool(SYSTEM_SENDER_RE.search(sender or "") or SYSTEM_SUBJECT_RE.search(subject or ""))


def subject_rank(subject: str) -> int:
    """Subject helps ordering only.  Validity is determined from Excel columns/data."""
    s = norm(subject)
    score = 10 * sum(1 for key in SUBJECT_KEYS if key and key in s)
    if "21vek" in s or "21 век" in s:
        score += 5
    if "реализац" in s or "продаж" in s:
        score += 5
    return score


def col(headers, *names):
    h = [norm(x) for x in headers]
    for n in names:
        k = norm(n)
        for i, x in enumerate(h):
            if x == k or x.startswith(k):
                return i
    return -1


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for f in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], f).date()
        except Exception:
            pass
    return None


def parse_xlsx(content, filename):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    target = None
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 30) + 1):
            headers = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            sku = col(headers, "Артикул", "SKU")
            qty = col(headers, "Количество", "Кол-во", "Кол.")
            if sku >= 0 and qty >= 0:
                target = (
                    ws,
                    r,
                    headers,
                    sku,
                    qty,
                    col(headers, "Дата", "Дата документа", "Период"),
                    col(headers, "Номенклатура", "Товар", "Наименование"),
                    col(headers, "Документ", "Номер документа", "Регистратор"),
                    col(headers, "Сумма", "Выручка", "Сумма с НДС"),
                    col(headers, "Контрагент", "Клиент"),
                )
                break
        if target:
            break
    if not target:
        raise RuntimeError("Не найдена таблица: нужны колонки Артикул и Количество")

    ws, hr, headers, csku, cqty, cdt, cprod, cdoc, csum, cclient = target
    acc = {}
    for r in range(hr + 1, ws.max_row + 1):
        sku = str(ws.cell(r, csku + 1).value or "").strip().lstrip("'")
        if not ARTICLE_RE.match(sku) or re.match(r"^900(?:/|$)", sku, re.I):
            continue
        if cclient >= 0 and "триовист" not in norm(ws.cell(r, cclient + 1).value):
            continue
        qty = dec(ws.cell(r, cqty + 1).value)
        if qty == 0:
            continue
        dt = parse_date(ws.cell(r, cdt + 1).value) if cdt >= 0 else date.today()
        dt = dt or date.today()
        doc = str(ws.cell(r, cdoc + 1).value or "").strip() if cdoc >= 0 else ""
        doc = doc or "без документа"
        product = str(ws.cell(r, cprod + 1).value or "").strip() if cprod >= 0 else ""
        amount = dec(ws.cell(r, csum + 1).value) if csum >= 0 else Decimal("0")
        key = (dt.isoformat(), doc, sku)
        x = acc.setdefault(
            key,
            {
                "shipment_date": dt.isoformat(),
                "document_no": doc,
                "sku": sku,
                "product": product,
                "qty": Decimal("0"),
                "amount": Decimal("0"),
                "source_file": filename,
            },
        )
        x["qty"] += qty
        x["amount"] += amount

    return [
        {
            **x,
            "qty": format(x["qty"], "f"),
            "amount": format(x["amount"].quantize(Decimal("0.01")), "f"),
        }
        for x in acc.values()
    ]


def upsert(rows):
    """Upsert only.  Never deletes a previous shipment snapshot."""
    if not rows:
        return 0
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": "Bearer " + SUPABASE_KEY,
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    for i in range(0, len(rows), 500):
        r = requests.post(
            SUPABASE_URL + "/rest/v1/triovist_shipments?on_conflict=shipment_date,document_no,sku",
            headers=headers,
            json=rows[i : i + 500],
            timeout=120,
        )
        if r.status_code not in (200, 201, 204):
            raise RuntimeError(f"Supabase {r.status_code}: {r.text}")
    return len(rows)


def fetch_message_headers_batch(mail, message_ids):
    """Fetch Subject/Date for a batch in one IMAP round-trip."""
    if not message_ids:
        return []
    message_set = b",".join(message_ids)
    status, data = mail.fetch(message_set, "(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE FROM)])")
    if status != "OK":
        raise RuntimeError("Gmail не вернул пакет заголовков")
    requested = {bytes(x) for x in message_ids}
    result = []
    for item in data or []:
        if not (isinstance(item, tuple) and len(item) > 1 and isinstance(item[1], (bytes, bytearray))):
            continue
        meta_raw = item[0] if isinstance(item[0], (bytes, bytearray)) else b""
        m = re.match(rb"\s*(\d+)", bytes(meta_raw))
        if not m:
            continue
        message_id = m.group(1)
        if message_id not in requested:
            continue
        raw = bytes(item[1])
        if not raw:
            continue
        hdr = email.message_from_bytes(raw)
        result.append({"id": message_id, "subject": decode(hdr.get("Subject")), "sender": decode(hdr.get("From")), "sent": msg_date(hdr.get("Date"))})
    return result


def recent_excel_candidates(mail, ids):
    """Batch-scan headers, reject system mail and rank recent Excel-bearing candidates."""
    candidates = []
    started = time.monotonic()
    total_batches = max(1, (len(ids) + HEADER_BATCH_SIZE - 1) // HEADER_BATCH_SIZE)
    for batch_no, start in enumerate(range(0, len(ids), HEADER_BATCH_SIZE), 1):
        if time.monotonic() - started > HEADER_SCAN_DEADLINE:
            raise RuntimeError(
                f"Gmail слишком долго отдаёт заголовки: превышен лимит {HEADER_SCAN_DEADLINE} сек. "
                "Отгрузки в CRM не изменялись."
            )
        batch = ids[start : start + HEADER_BATCH_SIZE]
        log(f"  Заголовки {batch_no}/{total_batches}: {len(batch)} писем…")
        metas = fetch_message_headers_batch(mail, batch)
        for meta in metas:
            if is_system_notice(meta.get("subject", ""), meta.get("sender", "")):
                log("    пропуск системного письма: " + str(meta.get("subject") or "")[:120])
                continue
            meta["rank"] = subject_rank(meta.get("subject", ""))
            candidates.append(meta)
    candidates.sort(key=lambda x: (x.get("rank", 0), x["sent"]), reverse=True)
    log(f"Excel-писем после исключения системных уведомлений: {len(candidates)}.")
    return candidates[:MAX_FULL_MESSAGES]


def search_recent_excel_ids(mail):
    """Use Gmail-native attachment search; fall back safely if X-GM-RAW is unavailable."""
    raw_query = (
        f"newer_than:{LOOKBACK}d has:attachment {{filename:xlsx filename:xlsm}} "
        "-from:notifications@github.com -from:noreply@github.com"
    )
    try:
        st, data = mail.search(None, "X-GM-RAW", '"' + raw_query + '"')
        if st == "OK":
            ids = (data[0] or b"").split()
            log(f"Gmail: найдено Excel-писем за последние {LOOKBACK} дн.: {len(ids)}.")
            return ids
        log("  ⚠️ Gmail X-GM-RAW вернул статус " + str(st) + "; использую безопасный fallback.")
    except Exception as e:
        log("  ⚠️ Gmail X-GM-RAW недоступен: " + str(e) + "; использую безопасный fallback.")

    since = (date.today() - timedelta(days=LOOKBACK)).strftime("%d-%b-%Y")
    st, data = mail.search(None, f"(SINCE {since})")
    if st != "OK":
        raise RuntimeError("Не удалось получить список писем")
    ids = (data[0] or b"").split()
    log(f"Fallback: писем за последние {LOOKBACK} дн.: {len(ids)}. Будут проверены заголовки и вложения.")
    return ids


def _full_message_bytes(raw_data):
    if not raw_data:
        return None
    for item in raw_data:
        if isinstance(item, tuple) and len(item) > 1 and isinstance(item[1], (bytes, bytearray)):
            return bytes(item[1])
    return None


def main():
    mail = connect_mail()
    try:
        select_box(mail)
        ids = search_recent_excel_ids(mail)
        if not ids:
            log("ℹ️ Новых Excel-вложений за период нет. Это нормальный запуск; данные отгрузок не менялись.")
            return
        log(f"Заголовки получаю пакетами по {HEADER_BATCH_SIZE}.")
        candidates = recent_excel_candidates(mail, ids)
        if not candidates:
            log("ℹ️ После исключения системных уведомлений подходящих Excel-писем нет. Данные не менялись.")
            return

        errors = []
        merged = {}
        valid_files = 0
        zero_reports = 0
        # Candidates are ranked/newest first.  If the same document/SKU appears in
        # several cumulative reports, keep the newest copy instead of summing it twice.
        for idx, meta in enumerate(candidates, 1):
            mid, sent, subject = meta["id"], meta["sent"], meta["subject"]
            sender = meta.get("sender", "")
            log(f"  Проверяю письмо {idx}/{len(candidates)}: {sent:%d.%m.%Y %H:%M} — {subject} — {sender}")
            try:
                st, raw_data = mail.fetch(mid, "(RFC822)")
            except (imaplib.IMAP4.abort, socket.timeout, TimeoutError, OSError) as exc:
                errors.append(f"{sent:%d.%m %H:%M}: IMAP {exc}")
                continue
            if st != "OK":
                errors.append(f"{sent:%d.%m %H:%M}: Gmail не вернул полное письмо")
                continue
            raw = _full_message_bytes(raw_data)
            if not raw:
                errors.append(f"{sent:%d.%m %H:%M}: пустое тело письма")
                continue
            msg = email.message_from_bytes(raw)
            if is_system_notice(decode(msg.get("Subject")), decode(msg.get("From"))):
                log("    системное письмо GitHub пропущено после полной проверки")
                continue
            excel = attachments(msg)
            if not excel:
                # X-GM-RAW normally prevents this, but a fallback search may include it.
                log("    Excel-вложений нет — пропуск")
                continue
            for fn, content in excel:
                try:
                    rows = parse_xlsx(content, fn)
                except Exception as e:
                    errors.append(f"{sent:%d.%m %H:%M} · {fn}: {e}")
                    log("    не отчёт отгрузок: " + errors[-1])
                    continue
                valid_files += 1
                if not rows:
                    zero_reports += 1
                    log(f"    ✅ Структура отчёта распознана: {fn}, но новых строк Триовист/разрешённых SKU нет.")
                    continue
                for row in rows:
                    key = (str(row.get("shipment_date")), str(row.get("document_no")), str(row.get("sku")))
                    merged.setdefault(key, row)
                log(f"    ✅ Валидный отчёт 1С: {fn}, строк Триовист: {len(rows)}")

        if merged:
            rows = list(merged.values())
            n = upsert(rows)
            log(
                f"✅ Отгрузки Триовист обновлены: {n} уникальных строк из {valid_files} валидных Excel. "
                "Повторы одного документа/SKU между отчётами не суммировались."
            )
            return
        if valid_files or zero_reports:
            log("ℹ️ Валидный отчёт 1С найден, но новых строк для загрузки нет. Это нормальный запуск; данные не менялись.")
            return

        detail = "; ".join(errors[-8:]) if errors else "валидный Excel-отчёт 1С не найден"
        log("ℹ️ Excel-вложения были, но ни одно не является отчётом отгрузок Триовист. Данные не менялись. " + detail)
        return
    finally:
        try:
            mail.logout()
        except Exception:
            pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log("ОШИБКА: " + str(e))
        sys.exit(1)
