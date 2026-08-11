#!/usr/bin/env python3
"""
Импорт свежего остатка склада Витебск из Gmail в Supabase (stock_balances).

v22.7.31.3 — восстановление автоматического обновления Витебска + реестр первого положительного остатка:
- Gmail жёстко закреплён на imap.gmail.com:993, есть ограниченные повторы;
- поиск идёт по всей почте Gmail (All Mail), а не только INBOX;
- сначала ищутся свежие письма с Excel-вложениями, тема используется только как приоритет;
- проверяются несколько свежих Excel-вложений, валидность определяет структура отчёта;
- дата отчёта берётся как самая поздняя корректная дата в шапке (а не первая дата периода);
- если явной даты в Excel нет, безопасно используется дата письма И она записывается в report_date;
- старый/пустой/невалидный отчёт никогда не перезаписывает склад;
- при ошибке записи выполняется попытка отката к предыдущему снимку;
- новинка фиксируется только при первом в истории положительном свободном остатке Витебска > 0 через постоянный реестр triovist_first_positive_stock;
- повторный приход уже известного положительного SKU новинкой не становится.
"""
from __future__ import annotations

import email
import imaplib
import io
import os
import re
import socket
import sys
import time
from datetime import date, datetime, timedelta, timezone
from email.header import decode_header
from email.utils import parsedate_to_datetime

import openpyxl
import requests

VERSION = "v22.7.31.3"
GMAIL_HOST = "imap.gmail.com"
GMAIL_PORT = 993
ARTICLE_RE = re.compile(r"^\d+(?:/\d+){1,8}$")


def _env_text(name: str, default: str = "") -> str:
    value = os.getenv(name)
    if value is None:
        return default
    value = str(value).strip()
    return value or default


def _safe_int(name: str, default: int, lo: int, hi: int) -> int:
    try:
        n = int(_env_text(name, str(default)))
    except Exception:
        n = default
    return max(lo, min(hi, n))


# Dedicated Gmail importer: shared/blank IMAP secrets must never redirect it.
IMAP_HOST = GMAIL_HOST
IMAP_PORT = GMAIL_PORT
IMAP_USER = os.environ["IMAP_USER"].strip()
IMAP_PASS = "".join(os.environ["IMAP_PASS"].split()).replace("\xa0", "")
SUPABASE_URL = os.environ["SUPABASE_URL"].strip().rstrip("/")
SUPABASE_KEY = os.environ["SUPABASE_KEY"].strip()

WAREHOUSE = _env_text("STOCK_WAREHOUSE", "Витебск")
LOOKBACK_DAYS = _safe_int("STOCK_LOOKBACK_DAYS", 10, 3, 30)
MAX_REPORT_AGE_DAYS = _safe_int("STOCK_MAX_AGE_DAYS", 3, 1, 14)
MAIL_TIMEOUT = _safe_int("STOCK_IMAP_TIMEOUT_SECONDS", 45, 15, 120)
CONNECT_ATTEMPTS = _safe_int("STOCK_IMAP_ATTEMPTS", 3, 1, 5)
CONNECT_RETRY_SECONDS = _safe_int("STOCK_IMAP_RETRY_SECONDS", 6, 1, 30)
HEADER_BATCH_SIZE = _safe_int("STOCK_IMAP_HEADER_BATCH", 75, 20, 250)
MAX_FULL_MESSAGES = _safe_int("STOCK_MAX_EMAIL_CANDIDATES", 30, 5, 60)
SUBJECT_KEYS = tuple(
    x.strip().lower()
    for x in _env_text("STOCK_SUBJECT_KEYS", "остатк,crm,витебск").split(",")
    if x.strip()
)


def log(msg):
    print(msg, flush=True)


def norm(v):
    return re.sub(r"\s+", " ", str(v or "").strip().lower().replace("ё", "е"))


def num(value):
    if value is None or value == "":
        return None
    try:
        if isinstance(value, str):
            value = value.replace("\xa0", "").replace(" ", "").replace(",", ".")
        return float(value)
    except (TypeError, ValueError):
        return None


def decode(v):
    if not v:
        return ""
    out = []
    for part, enc in decode_header(v):
        if not isinstance(part, bytes):
            out.append(part)
            continue
        encoding = enc or "utf-8"
        if str(encoding).lower() == "unknown-8bit":
            encoding = "utf-8"
        try:
            out.append(part.decode(encoding, "replace"))
        except (LookupError, UnicodeDecodeError):
            out.append(part.decode("utf-8", "replace"))
    return "".join(out)


def msg_date(v):
    try:
        d = parsedate_to_datetime(v)
        if d.tzinfo is None:
            d = d.replace(tzinfo=timezone.utc)
        return d.astimezone(timezone.utc)
    except Exception:
        return datetime.min.replace(tzinfo=timezone.utc)


def connect_mail() -> imaplib.IMAP4_SSL:
    log(f"{VERSION}: подключаюсь к Gmail {IMAP_HOST}:{IMAP_PORT}")
    last_exc = None
    for attempt in range(1, CONNECT_ATTEMPTS + 1):
        mail = None
        try:
            log(f"  IMAP попытка {attempt}/{CONNECT_ATTEMPTS}…")
            mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT, timeout=MAIL_TIMEOUT)
            mail.login(IMAP_USER, IMAP_PASS)
            try:
                mail.sock.settimeout(MAIL_TIMEOUT)
            except Exception:
                pass
            log("  ✅ Gmail IMAP подключён")
            return mail
        except imaplib.IMAP4.error as exc:
            last_exc = exc
            try:
                if mail is not None:
                    mail.logout()
            except Exception:
                pass
            if re.search(r"auth|authentication|credentials|login|invalid", str(exc), re.I):
                raise RuntimeError(f"Gmail отклонил вход IMAP: {exc}") from exc
        except (imaplib.IMAP4.abort, socket.timeout, TimeoutError, ConnectionError, OSError) as exc:
            last_exc = exc
            try:
                if mail is not None:
                    mail.logout()
            except Exception:
                pass
        if attempt < CONNECT_ATTEMPTS:
            log(f"  ⚠️ IMAP временно недоступен: {last_exc}. Повтор через {CONNECT_RETRY_SECONDS} сек")
            time.sleep(CONNECT_RETRY_SECONDS)
    raise RuntimeError(f"Не удалось подключиться к Gmail после {CONNECT_ATTEMPTS} попыток: {last_exc}")


def select_box(mail):
    """Gmail All Mail first, INBOX only as fallback."""
    try:
        st, boxes = mail.list()
        if st == "OK":
            for raw in boxes or []:
                if b"\\All" not in raw:
                    continue
                m = re.search(br'\)\s+"[^"]*"\s+(.+)$', raw)
                if not m:
                    continue
                token = m.group(1).strip()
                if token.startswith(b'"') and token.endswith(b'"'):
                    token = token[1:-1]
                mailbox = token.decode("ascii", "ignore")
                if mailbox:
                    arg = f'"{mailbox}"' if " " in mailbox else mailbox
                    if mail.select(arg)[0] == "OK":
                        log("Ищу остатки во всей почте: " + mailbox)
                        return mailbox
    except Exception as exc:
        log("  ⚠️ All Mail не открылся: " + str(exc))
    if mail.select("INBOX")[0] != "OK":
        raise RuntimeError("Не удалось открыть Gmail All Mail или INBOX")
    log("Ищу остатки в INBOX")
    return "INBOX"


def read_report_date(ws):
    """Return the latest plausible date found in the report header.

    Old code returned the FIRST dd.mm.yyyy it saw.  If 1C printed a period such
    as 01.08.2026–10.08.2026, CRM stored 01.08 and immediately considered the
    stock stale.  For a point-in-time stock report the latest date in the header
    is the safe relevant boundary.
    """
    dates = []
    today = date.today()
    for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), max_col=min(20, ws.max_column)):
        for cell in row:
            value = cell.value
            if isinstance(value, datetime):
                dates.append(value.date())
                continue
            if isinstance(value, date):
                dates.append(value)
                continue
            if value in (None, ""):
                continue
            text = str(value)
            for d, mth, y in re.findall(r"(?<!\d)(\d{1,2})[./-](\d{1,2})[./-](\d{4})(?!\d)", text):
                try:
                    dates.append(date(int(y), int(mth), int(d)))
                except ValueError:
                    pass
    plausible = [d for d in dates if d <= today + timedelta(days=1)]
    return max(plausible) if plausible else None


def find_header_row(ws):
    for r in range(1, min(35, ws.max_row) + 1):
        for c in range(1, min(6, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if v and norm(v) in {"артикул", "sku"}:
                return r, c
    raise RuntimeError("Не найдена шапка с колонкой «Артикул»")


def col_index(ws, header_row, *names):
    wanted = [norm(n) for n in names]
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if not v:
            continue
        low = norm(v)
        if any(low == n or low.startswith(n) for n in wanted):
            return c
    return None


def parse_stock(xlsx_bytes, fallback_date=None):
    """Parse a stock report. Returns (rows, explicit_report_date)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    last_error = None
    for ws in wb.worksheets:
        try:
            report_date = read_report_date(ws)
            hr, _ = find_header_row(ws)
            c_sku = col_index(ws, hr, "артикул", "sku")
            c_name = col_index(ws, hr, "номенклатура", "товар", "наименование")
            c_unit = col_index(ws, hr, "ед")
            c_onh = col_index(ws, hr, "в наличии", "наличи")
            c_res = col_index(ws, hr, "в резерве", "резерв")
            c_avail = col_index(ws, hr, "доступно", "свободно")
            if c_sku is None or c_avail is None:
                raise RuntimeError("нет обязательных колонок Артикул + Доступно")

            effective = report_date or fallback_date
            rows = []
            seen = set()
            for r in range(hr + 1, ws.max_row + 1):
                sku = str(ws.cell(row=r, column=c_sku).value or "").strip().lstrip("'")
                if not ARTICLE_RE.match(sku) or sku in seen:
                    continue
                seen.add(sku)
                name = ws.cell(row=r, column=c_name).value if c_name else None
                avail = num(ws.cell(row=r, column=c_avail).value)
                onh = num(ws.cell(row=r, column=c_onh).value) if c_onh else None
                res = num(ws.cell(row=r, column=c_res).value) if c_res else None
                unit = ws.cell(row=r, column=c_unit).value if c_unit else None
                rows.append({
                    "sku": sku,
                    "product": str(name).strip() if name else None,
                    "warehouse": WAREHOUSE,
                    "unit": str(unit).strip() if unit not in (None, "") else None,
                    "qty_onhand": onh,
                    "qty_reserve": res,
                    "qty_avail": avail,
                    "report_date": effective.isoformat() if effective else None,
                })
            if not rows:
                raise RuntimeError("товарных строк с артикулами не найдено")
            return rows, report_date
        except Exception as exc:
            last_error = exc
            continue
    raise RuntimeError(f"Excel не распознан как отчёт остатков Витебска: {last_error}")


def attachments(msg):
    out = []
    for part in msg.walk():
        fn = decode(part.get_filename()).strip()
        if not fn or not fn.lower().endswith((".xlsx", ".xlsm")):
            continue
        payload = part.get_payload(decode=True)
        if payload:
            out.append((fn, payload))
    return out


SYSTEM_SENDER_RE = re.compile(r"(?:notifications|noreply)@github\.com|github-actions", re.I)
SYSTEM_SUBJECT_RE = re.compile(r"(?:\[resanta-crm/|run failed:|workflow run|github actions)", re.I)


def is_system_notice(subject, sender):
    return bool(SYSTEM_SENDER_RE.search(sender or "") or SYSTEM_SUBJECT_RE.search(subject or ""))


def subject_rank(subject):
    s = norm(subject)
    score = 0
    for key in SUBJECT_KEYS:
        if key and key in s:
            score += 10
    if "остат" in s:
        score += 20
    if "витебск" in s:
        score += 10
    if "1с" in s or "1c" in s:
        score += 5
    return score


def search_recent_excel_ids(mail):
    raw_query = (
        f"newer_than:{LOOKBACK_DAYS}d has:attachment {{filename:xlsx filename:xlsm}} "
        "-from:notifications@github.com -from:noreply@github.com"
    )
    try:
        st, data = mail.search(None, "X-GM-RAW", '"' + raw_query + '"')
        if st == "OK":
            ids = (data[0] or b"").split()
            log(f"Gmail: Excel-писем за последние {LOOKBACK_DAYS} дн.: {len(ids)}")
            return ids
    except Exception as exc:
        log("  ⚠️ X-GM-RAW недоступен: " + str(exc))
    since = (date.today() - timedelta(days=LOOKBACK_DAYS)).strftime("%d-%b-%Y")
    st, data = mail.search(None, f"(SINCE {since})")
    if st != "OK":
        raise RuntimeError("Не удалось получить список свежих писем")
    ids = (data[0] or b"").split()
    log(f"Fallback: писем за период: {len(ids)}")
    return ids


def fetch_headers_batch(mail, message_ids):
    if not message_ids:
        return []
    st, data = mail.fetch(b",".join(message_ids), "(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE FROM)])")
    if st != "OK":
        raise RuntimeError("Gmail не вернул пакет заголовков")
    requested = {bytes(x) for x in message_ids}
    out = []
    for item in data or []:
        if not (isinstance(item, tuple) and len(item) > 1 and isinstance(item[1], (bytes, bytearray))):
            continue
        meta = item[0] if isinstance(item[0], (bytes, bytearray)) else b""
        m = re.match(rb"\s*(\d+)", bytes(meta))
        if not m or m.group(1) not in requested:
            continue
        hdr = email.message_from_bytes(bytes(item[1]))
        out.append({
            "id": m.group(1),
            "subject": decode(hdr.get("Subject")),
            "sender": decode(hdr.get("From")),
            "sent": msg_date(hdr.get("Date")),
        })
    return out


def recent_candidates(mail, ids):
    rows = []
    total = max(1, (len(ids) + HEADER_BATCH_SIZE - 1) // HEADER_BATCH_SIZE)
    for no, start in enumerate(range(0, len(ids), HEADER_BATCH_SIZE), 1):
        batch = ids[start:start + HEADER_BATCH_SIZE]
        log(f"  Заголовки {no}/{total}: {len(batch)} писем…")
        for meta in fetch_headers_batch(mail, batch):
            if is_system_notice(meta["subject"], meta["sender"]):
                continue
            meta["rank"] = subject_rank(meta["subject"])
            rows.append(meta)
    # Newest first; subject only helps within similar recency and avoids random Excel first.
    rows.sort(key=lambda x: (x["sent"], x["rank"]), reverse=True)
    return rows[:MAX_FULL_MESSAGES]


def _full_message_bytes(raw_data):
    for item in raw_data or []:
        if isinstance(item, tuple) and len(item) > 1 and isinstance(item[1], (bytes, bytearray)):
            return bytes(item[1])
    return None


def _rest_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }


def load_previous_stock():
    """Read complete known stock columns before replacement; also used by novelty registry."""
    select = "sku,product,warehouse,unit,qty_onhand,qty_reserve,qty_avail,report_date"
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/stock_balances?warehouse=eq.{WAREHOUSE}&select={select}",
        headers=_rest_headers(), timeout=120,
    )
    if r.status_code != 200:
        log(f"⚠️ Не удалось прочитать предыдущий остаток: {r.status_code}")
        return {}, []
    rows = r.json() or []
    return {str(x.get("sku") or "").strip(): x for x in rows if x.get("sku")}, rows


def _insert_stock_rows(rows):
    headers = {**_rest_headers(), "Prefer": "return=minimal"}
    for i in range(0, len(rows), 500):
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/stock_balances",
            headers=headers, json=rows[i:i + 500], timeout=120,
        )
        if r.status_code not in (200, 201, 204):
            raise RuntimeError(f"Supabase insert {r.status_code}: {r.text[:600]}")
        log(f"  Записано {min(i + 500, len(rows))} / {len(rows)}")


def replace_stock(rows, previous_rows):
    if not rows:
        raise RuntimeError("Пустой разбор: склад НЕ изменён")
    headers = _rest_headers()
    r = requests.delete(
        f"{SUPABASE_URL}/rest/v1/stock_balances?warehouse=eq.{WAREHOUSE}",
        headers=headers, timeout=120,
    )
    if r.status_code not in (200, 204):
        raise RuntimeError(f"Не удалось очистить старый снимок: {r.status_code} {r.text[:600]}")
    log(f"  Старый снимок склада «{WAREHOUSE}» удалён")
    try:
        _insert_stock_rows(rows)
    except Exception as exc:
        log("  ⛔ Запись нового снимка не завершилась. Пытаюсь восстановить предыдущий снимок…")
        try:
            requests.delete(
                f"{SUPABASE_URL}/rest/v1/stock_balances?warehouse=eq.{WAREHOUSE}",
                headers=headers, timeout=120,
            )
            if previous_rows:
                _insert_stock_rows(previous_rows)
                log("  ✅ Предыдущий снимок Витебска восстановлен")
            else:
                log("  ⚠️ Предыдущего снимка для восстановления не было")
        except Exception as rollback_exc:
            raise RuntimeError(f"Ошибка записи: {exc}; также не удался откат: {rollback_exc}") from exc
        raise


def save_first_positive_stock(rows):
    """Persist the first-ever positive free stock event for Vitebsk.

    The SQL migration pre-seeds already-known positive/history SKU as legacy.
    Therefore an INSERT with on_conflict=sku is enough: once an SKU has a
    positive event in the registry, a later zero -> positive return cannot
    create a second novelty.
    """
    today = date.today().isoformat()
    out = []
    for x in rows:
        sku = str(x.get("sku") or "").strip()
        if not sku or re.match(r"^900(?:/|$)", sku, re.I):
            continue
        avail = float(x.get("qty_avail") or 0)
        if avail <= 0:
            continue
        out.append({
            "sku": sku,
            "product": x.get("product"),
            "first_positive_date": x.get("report_date") or today,
            "first_positive_qty": avail,
            "source": "stock_balances",
            "is_legacy": False,
            "updated_at": datetime.now().isoformat(),
        })
    if not out:
        return
    headers = {**_rest_headers(), "Prefer": "resolution=ignore-duplicates,return=minimal"}
    inserted = 0
    for i in range(0, len(out), 500):
        batch = out[i:i + 500]
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/triovist_first_positive_stock?on_conflict=sku",
            headers=headers, json=batch, timeout=120,
        )
        if r.status_code not in (200, 201, 204):
            # The warehouse snapshot is already safely replaced at this point;
            # registry failure must be visible but must not roll back stock data.
            log(f"⚠️ Реестр первого положительного остатка не сохранён: {r.status_code} {r.text[:300]}")
            return
        inserted += len(batch)
    log(f"🆕 Проверено положительных SKU для реестра: {inserted}")


def choose_latest_valid_report(mail, candidates):
    errors = []
    for idx, meta in enumerate(candidates, 1):
        mid, sent = meta["id"], meta["sent"]
        log(f"  Проверяю письмо {idx}/{len(candidates)}: {sent:%d.%m.%Y %H:%M} — {meta['subject']}")
        try:
            st, raw_data = mail.fetch(mid, "(RFC822)")
            if st != "OK":
                errors.append("Gmail не вернул письмо")
                continue
            raw = _full_message_bytes(raw_data)
            if not raw:
                errors.append("пустое тело письма")
                continue
            msg = email.message_from_bytes(raw)
            if is_system_notice(decode(msg.get("Subject")), decode(msg.get("From"))):
                continue
            excels = attachments(msg)
            if not excels:
                continue
            for fn, content in excels:
                try:
                    fallback = sent.date()
                    rows, explicit_date = parse_stock(content, fallback_date=fallback)
                except Exception as exc:
                    errors.append(f"{fn}: {exc}")
                    continue
                effective = explicit_date or fallback
                age = (date.today() - effective).days
                if age < 0:
                    errors.append(f"{fn}: дата отчёта в будущем {effective:%d.%m.%Y}")
                    continue
                if age > MAX_REPORT_AGE_DAYS:
                    errors.append(f"{fn}: отчёт {effective:%d.%m.%Y} старше {MAX_REPORT_AGE_DAYS} дн.")
                    continue
                # Ensure every row has the chosen date even when Excel had none.
                for row in rows:
                    row["report_date"] = effective.isoformat()
                log(
                    f"    ✅ Свежий отчёт: {fn}; дата {effective:%d.%m.%Y}; "
                    f"{len(rows)} SKU; источник даты: {'Excel' if explicit_date else 'дата письма'}"
                )
                return rows, effective, fn, sent
        except (imaplib.IMAP4.abort, socket.timeout, TimeoutError, OSError) as exc:
            errors.append(f"IMAP: {exc}")
            continue
    detail = "; ".join(errors[-10:]) if errors else "валидный свежий Excel не найден"
    raise RuntimeError(
        "За выбранный период не найден свежий валидный отчёт остатков Витебска. "
        "Старый склад НЕ изменён. Последние проверки: " + detail
    )


def main():
    mail = connect_mail()
    try:
        select_box(mail)
        ids = search_recent_excel_ids(mail)
        if not ids:
            raise RuntimeError(
                f"За последние {LOOKBACK_DAYS} дн. нет Excel-писем. Старый склад НЕ изменён."
            )
        candidates = recent_candidates(mail, ids)
        if not candidates:
            raise RuntimeError("Подходящих писем после фильтрации нет. Старый склад НЕ изменён.")

        rows, report_date, filename, sent = choose_latest_valid_report(mail, candidates)
        total_avail = sum(r["qty_avail"] or 0 for r in rows)
        log(f"  Суммарно доступно: {total_avail:,.0f} шт по {len(rows)} SKU")

        previous_map, previous_rows = load_previous_stock()
        replace_stock(rows, previous_rows)
        save_first_positive_stock(rows)
        log(
            f"✅ Витебск обновлён: report_date={report_date.isoformat()}, "
            f"письмо={sent:%d.%m.%Y %H:%M}, файл={filename}"
        )
    finally:
        try:
            mail.logout()
        except Exception:
            pass


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log("ОШИБКА: " + str(exc))
        sys.exit(1)
